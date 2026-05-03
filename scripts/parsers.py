#!/usr/bin/env python3
"""
Raito Dashboard — Data Parsers
All file parsing functions: Icedream, Ma'ayan, Karfree, Distributor Stock, Production.
"""
from __future__ import annotations

import re
import math
import logging
import pandas as pd
from openpyxl import load_workbook
from config import (
    DATA_DIR, OUTPUT_DIR, classify_product, extract_units_per_carton,
    MONTH_ORDER, extract_customer_name, INTERNAL_ACCOUNTS,
)
from pricing_engine import get_b2b_price_safe, get_production_cost, get_customer_price
from registry import validate_sku, PRODUCTS

_log = logging.getLogger(__name__)

# ── Optional DB Resolver (Phase 2: ID-based resolution) ─────────────────
_resolver_instance = None
_resolver_init_attempted = False


def _get_resolver():
    """Try to initialize the DB entity resolver (singleton).

    Returns an EntityResolver if DB is reachable, or None.
    Safe to call repeatedly — only attempts connection once per session.
    """
    global _resolver_instance, _resolver_init_attempted
    if _resolver_init_attempted:
        return _resolver_instance
    _resolver_init_attempted = True
    try:
        from db.resolvers import EntityResolver
        _resolver_instance = EntityResolver()
        _log.info("DB resolver initialized — ID-based resolution enabled")
        print("  ✓ DB resolver connected — ID enrichment active")
    except Exception as e:
        _log.info(f"DB resolver not available ({e}) — string-only mode")
        _resolver_instance = None
    return _resolver_instance


# Track unknown SKUs seen during this session (warn once per SKU)
_unknown_skus_warned: set = set()

def _validated_product(product: str) -> str | None:
    """Validate a classified product SKU against the registry.

    Returns the product key if valid, or None if unknown (logs a warning).
    """
    if product is None:
        return None
    if product in PRODUCTS:
        return product
    if product not in _unknown_skus_warned:
        _unknown_skus_warned.add(product)
        _log.warning(
            "Unregistered SKU '%s' detected during data ingestion — "
            "register it in registry.PRODUCTS before use", product
        )
        print(f"  ⚠ WARNING: Unregistered SKU '{product}' — add to registry.py")
    return product  # Pass through so data isn't silently lost


# ── Month Detection ─────────────────────────────────────────────────────

def detect_month_from_sheet(ws):
    """Detect month from sheet name and date range row (Row 3)."""
    title = ws.title or ''
    month_keywords = {
        'נובמבר': 'November 2025', 'November': 'November 2025', 'NOVEMBER': 'November 2025',
        'דצמבר': 'December 2025', 'December': 'December 2025', 'DECEMBER': 'December 2025',
        'ינואר': 'January 2026', 'January': 'January 2026', 'JANUARY': 'January 2026',
        'פברואר': 'February 2026', 'February': 'February 2026', 'FEBRUARY': 'February 2026',
        'מרץ': 'March 2026', 'March': 'March 2026', 'MARCH': 'March 2026',
    }
    for keyword, month in month_keywords.items():
        if keyword in title:
            return month

    for col in range(1, 7):
        val = ws.cell(row=3, column=col).value
        if val is None:
            continue
        val = str(val)
        date_match = re.search(r'(\d{1,2})/(\d{1,2})/(\d{4})', val)
        if date_match:
            day, month_num, year = date_match.groups()
            month_num = int(month_num)
            year = int(year)
            month_map = {
                (11, 2025): 'November 2025',
                (12, 2025): 'December 2025',
                (1, 2026): 'January 2026', (2, 2026): 'February 2026',
                (3, 2026): 'March 2026', (4, 2026): 'April 2026',
            }
            return month_map.get((month_num, year), f'{month_num}/{year}')

    for row in range(1, 5):
        for col in range(1, 6):
            val = ws.cell(row=row, column=col).value
            if val is None:
                continue
            val = str(val)
            for keyword, month in month_keywords.items():
                if keyword in val:
                    return month
    return None


# ── Icedreams Parser ─────────────────────────────────────────────────────

def parse_icedreams_file(filepath):
    """Parse a single Icedreams monthly report."""
    wb = load_workbook(filepath)
    ws = wb[wb.sheetnames[0]]
    month = detect_month_from_sheet(ws)
    if not month:
        month = ws.title if ws.title else 'Unknown'

    data = {'month': month, 'by_customer': {}, 'totals': {}}
    pending_products = []

    for row_idx in range(1, ws.max_row + 1):
        cell_a = ws.cell(row=row_idx, column=1)
        is_bold = cell_a.font.bold if cell_a.font else False
        item_name = ws.cell(row=row_idx, column=4).value
        sales_val = ws.cell(row=row_idx, column=5).value
        quantity = ws.cell(row=row_idx, column=6).value

        if is_bold and cell_a.value:
            s = str(cell_a.value).strip()
            if 'סה"כ' in s and 'לדו' not in s and 'מס' not in s:
                cell_b = ws.cell(row=row_idx, column=2)
                customer_name = str(cell_b.value).strip() if cell_b.value else s.replace('סה"כ', '').strip()
                if customer_name:
                    if customer_name not in data['by_customer']:
                        data['by_customer'][customer_name] = {}
                    for pp in pending_products:
                        p, u, v, c = pp
                        if p not in data['by_customer'][customer_name]:
                            data['by_customer'][customer_name][p] = {'units': 0, 'value': 0, 'cartons': 0}
                        data['by_customer'][customer_name][p]['units'] += u
                        data['by_customer'][customer_name][p]['value'] += v
                        data['by_customer'][customer_name][p]['cartons'] += c
                    pending_products = []

        if item_name and quantity is not None:
            product = _validated_product(classify_product(item_name))
            if product:
                upc = extract_units_per_carton(item_name)
                raw_qty = float(quantity)
                # Negative = sales, Positive = returns (subtract)
                sign = -1 if raw_qty < 0 else 1  # sales are negative in report
                cartons = abs(raw_qty)
                units = round(cartons * upc) * sign * -1  # flip: negative→positive sales
                value = float(sales_val) * sign * -1 if sales_val else 0

                if product not in data['totals']:
                    data['totals'][product] = {'units': 0, 'value': 0, 'cartons': 0}
                data['totals'][product]['units'] += units
                data['totals'][product]['value'] += value
                data['totals'][product]['cartons'] += cartons * sign * -1
                pending_products.append((product, units, value, cartons * sign * -1))

    wb.close()
    return data


def _parse_icedream_weekly_xlsx(filepath):
    """Parse a single-week Icedream .xlsx file (Format B weekly).

    Layout (e.g. week13.xlsx):
      R1: _, _, _, 'שבוע 13 2026'   (week identifier)
      R2: 'שם חשבון', 'שם פריט', 'כמות', 'בש"ח'  (headers)
      R3+: data rows — account in col 1 (only on first product row),
           product in col 2, cartons in col 3 (negative=sales), revenue in col 4.
           'סה"כ' in col 2 = customer summary row (skip).

    Returns same structure as parse_icedreams_file():
      {'month': str, 'by_customer': {name: {product: {units, value, cartons}}}, 'totals': {}}
    """
    import os
    wb = load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # --- Detect week number — scan first 5 rows, all columns ---
    week_num = None
    for r in range(1, 6):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if val and 'שבוע' in str(val):
                m = re.search(r'(\d+)', str(val))
                if m:
                    week_num = int(m.group(1))
                break
        if week_num:
            break

    # Map week to month (from central registry)
    if week_num:
        from config import WEEK_TO_MONTH
        month = WEEK_TO_MONTH.get(week_num, f'W{week_num}')
    else:
        month = detect_month_from_sheet(ws) or 'Unknown'

    print(f"  [Icedream weekly] {os.path.basename(filepath)} → W{week_num} → {month}")

    data = {'month': month, 'by_customer': {}, 'totals': {}}
    current_account = None

    def _clean_account(s):
        """Strip branch decorations to get a clean account name."""
        s = re.sub(r'^\*+|\*+$', '', s).strip()
        s = re.sub(r'\*ת\.משלוח.*$', '', s).strip()
        s = re.sub(r'ת\.משלוח.*$', '', s).strip()
        s = re.sub(r'\*[^*]*$', '', s).strip()
        return s.strip()

    # --- Find header row and detect column order ---
    header_row = None
    qty_col = 3   # default: col 3 = כמות (qty), col 4 = כספי (revenue)
    rev_col = 4
    for r in range(1, 8):
        v = ws.cell(row=r, column=2).value
        if v and 'פריט' in str(v):
            header_row = r
            # Check if columns are swapped (כספי before כמות)
            for c in range(3, ws.max_column + 1):
                hdr = str(ws.cell(row=r, column=c).value or '')
                if 'כספי' in hdr or 'כסף' in hdr:
                    rev_col = c
                elif 'כמות' in hdr:
                    qty_col = c
            break

    data_start = (header_row + 1) if header_row else 3
    max_row = ws.max_row or 1000  # Google Sheets exports may have None max_row

    for row_idx in range(data_start, max_row + 1):
        acct_raw = ws.cell(row=row_idx, column=1).value

        # Update current account when column 1 has a value
        if acct_raw:
            cleaned = _clean_account(str(acct_raw).strip())
            if cleaned:
                current_account = cleaned

        if not current_account:
            continue

        product_raw = ws.cell(row=row_idx, column=2).value
        if not product_raw:
            continue
        product_str = str(product_raw).strip()

        # Skip summary rows
        if 'סה"כ' in product_str:
            continue

        # Strict Raito product filter
        if 'טורבו' not in product_str and 'דרים קייק' not in product_str:
            continue

        product = _validated_product(classify_product(product_str))
        if not product:
            continue

        upc = extract_units_per_carton(product_str)

        qty_raw = ws.cell(row=row_idx, column=qty_col).value
        rev_raw = ws.cell(row=row_idx, column=rev_col).value

        qty = float(qty_raw) if qty_raw not in (None, '', ' ') else 0.0
        rev = float(rev_raw) if rev_raw not in (None, '', ' ') else 0.0

        # Negative = sales, flip to positive
        cartons = abs(qty)
        units = round(cartons * upc)
        value = round(abs(rev), 2)

        if units == 0 and value == 0:
            continue

        # Skip summary rows that might appear as account names
        if current_account and 'סה"כ' in current_account:
            continue

        # Totals
        if product not in data['totals']:
            data['totals'][product] = {'units': 0, 'value': 0, 'cartons': 0}
        data['totals'][product]['units'] += units
        data['totals'][product]['value'] += value
        data['totals'][product]['cartons'] += cartons

        # By customer — keep raw per-branch name so SP tab sees per-branch rows.
        # Downstream CC/BO call extract_customer_name() at read time to aggregate.
        cust_name = current_account
        if cust_name not in data['by_customer']:
            data['by_customer'][cust_name] = {}
        if product not in data['by_customer'][cust_name]:
            data['by_customer'][cust_name][product] = {'units': 0, 'value': 0, 'cartons': 0}
        data['by_customer'][cust_name][product]['units'] += units
        data['by_customer'][cust_name][product]['value'] += value
        data['by_customer'][cust_name][product]['cartons'] += cartons

    wb.close()
    return data


def _is_icedream_weekly_xlsx(filepath):
    """Check if an xlsx file is a single-week Format B file (e.g. week13.xlsx).

    Detected by 'שבוע' in the first row of the first sheet.
    """
    try:
        wb = load_workbook(filepath, read_only=True)
        ws = wb[wb.sheetnames[0]]
        # Scan first 5 rows — week header may appear in row 1 (old format) or row 3 (W15+ format)
        # Only count as single-week if exactly 1 'שבוע' cell found (multi-week files have 2+)
        shavua_count = 0
        for r in range(1, 6):
            for c in range(1, min(13, ws.max_column + 1 if ws.max_column else 13)):
                val = ws.cell(row=r, column=c).value
                if val and 'שבוע' in str(val):
                    shavua_count += 1
        wb.close()
        return shavua_count == 1
    except Exception:
        pass
    return False


def _is_format_b_multiweek_xlsx(filepath):
    """Detect a multi-week Format B xlsx (e.g. converted sales_week_12.xlsx).

    Layout: row 1 = title, row 3 = week headers ('שבוע N'), row 4 = col labels.
    Distinguished from single-week files where 'שבוע' appears in row 1.
    """
    try:
        wb = load_workbook(filepath, read_only=True)
        ws = wb[wb.sheetnames[0]]
        # Row 1 must NOT have 'שבוע' (that would be single-week format)
        row1_has_shavua = any(
            ws.cell(row=1, column=c).value and 'שבוע' in str(ws.cell(row=1, column=c).value)
            for c in range(1, 7)
        )
        if row1_has_shavua:
            wb.close()
            return False
        # Row 3 must have 'שבוע' in at least 2 of cols 2–12
        # (multi-week files have multiple week columns; single-week W15 format has only one)
        row3_shavua_count = sum(
            1 for c in range(2, 13)
            if ws.cell(row=3, column=c).value and 'שבוע' in str(ws.cell(row=3, column=c).value)
        )
        wb.close()
        return row3_shavua_count >= 2
    except Exception:
        pass
    return False


def _format_b_read_rows(filepath):
    """Read a Format B file (either .xls or .xlsx) into a list of row lists.

    Tries openpyxl first for .xlsx, falls back to xlrd for .xls.
    Returns list of rows where each row is a list of cell values.
    """
    from pathlib import Path as _Path
    p = _Path(filepath)
    if p.suffix.lower() == '.xlsx':
        wb = load_workbook(p, read_only=True)
        ws = wb.active
        rows = [[c.value for c in row] for row in ws.iter_rows()]
        wb.close()
        return rows
    else:
        try:
            import xlrd as _xlrd
            wb = _xlrd.open_workbook(str(p))
            ws = wb.sheets()[0]
            return [[ws.cell_value(r, c) for c in range(ws.ncols)] for r in range(ws.nrows)]
        except ImportError:
            return []


def parse_format_b_xls(filepath):
    """Parse an Icedream Format B multi-week file (.xls or .xlsx).

    Supports both legacy .xls (via xlrd when available) and converted .xlsx
    (via openpyxl — no xlrd dependency required).

    Column layout (3-week file):
      0: Account name (only on first product row per customer)
      1: Product name (Hebrew SKU)
      2: W10 qty / 3: W10 revenue
      4: W11 qty / 5: W11 revenue
      6: W12 qty / 7: W12 revenue
      8: total qty / 9: total revenue

    Returns:
      {chain_en: {wk_idx: {product: {'units': int, 'value': float}}}}
      where wk_idx is 0-based (0=first week col, 1=second, etc.)

    Special rule: 'Oogiplatset' (עוגיפלצת) is only included for the LAST week
    column (W12), not for earlier weeks, per dashboard convention.
    """
    import re as _re_local

    try:
        all_rows = _format_b_read_rows(filepath)
    except Exception:
        return {}

    if len(all_rows) < 5:
        return {}

    n_cols = max(len(r) for r in all_rows)

    # Cols 0=account, 1=product, then pairs (qty,rev) per week, last 2 = totals
    data_start_col = 2
    week_cols = []
    for j in range(data_start_col, n_cols - 2, 2):
        week_cols.append((j, j + 1))

    def _clean(s):
        """Strip branch decorations to get a clean chain name."""
        s = _re_local.sub(r'^\*+|\*+$', '', s).strip()
        s = _re_local.sub(r'\*ת\.משלוח.*$', '', s).strip()
        s = _re_local.sub(r'ת\.משלוח.*$', '', s).strip()
        s = _re_local.sub(r'\*[^*]*$', '', s).strip()
        return s.strip()

    n_weeks = len(week_cols)
    customers = {}   # {chain_en: {wk_idx: {product: {'units', 'value'}}}}
    current_chain = None

    for row in all_rows[4:]:   # data starts at row index 4
        # Pad short rows
        while len(row) < n_cols:
            row.append(None)
        account_raw = str(row[0]).strip() if row[0] else ''
        product_raw = str(row[1]).strip() if row[1] else ''

        # Skip summary rows
        if product_raw == 'סה"כ' or account_raw == 'סה"כ':
            continue

        if account_raw:
            # Preserve raw per-branch name; CC/BO aggregate via extract_customer_name at read time.
            current_chain = _clean(account_raw)

        if not current_chain or not product_raw:
            continue

        # Strict Raito product filter
        if 'טורבו' not in product_raw and 'דרים קייק' not in product_raw:
            continue

        product = _validated_product(classify_product(product_raw))
        if not product:
            continue

        pack_size = extract_units_per_carton(product_raw)
        is_ugipletzet = 'Oogiplatset' in current_chain or 'עוגיפלצת' in current_chain

        if current_chain not in customers:
            customers[current_chain] = {i: {} for i in range(n_weeks)}

        for wk_i, (qi, vi) in enumerate(week_cols):
            # Oogiplatset rule: only include data from the LAST week column
            if is_ugipletzet and wk_i < n_weeks - 1:
                continue

            qty_raw = row[qi] if qi < len(row) and row[qi] != '' else 0
            rev_raw = row[vi] if vi < len(row) and row[vi] != '' else 0
            try:
                qty = float(qty_raw) if qty_raw not in ('', None) else 0.0
                rev = float(rev_raw) if rev_raw not in ('', None) else 0.0
            except (TypeError, ValueError):
                qty = rev = 0.0

            units = int(round(abs(qty) * pack_size))
            value = round(abs(rev), 2)

            if units > 0 or value > 0:
                if product not in customers[current_chain][wk_i]:
                    customers[current_chain][wk_i][product] = {'units': 0, 'value': 0.0}
                customers[current_chain][wk_i][product]['units'] += units
                customers[current_chain][wk_i][product]['value'] += value

    return customers


def parse_all_icedreams():
    """Parse all Icedreams files in the icedreams folder."""
    folder = DATA_DIR / 'icedreams'
    if not folder.exists():
        return {}
    results = {}
    for f in sorted(folder.glob('*.xlsx')):
        if f.name.startswith('~'):
            continue
        if 'stock' in f.name.lower():
            continue
        # Skip multi-week Format B xlsx files — handled by the supplement loop below
        if _is_format_b_multiweek_xlsx(f):
            continue
        # Route to weekly parser if single-week Format B xlsx
        if _is_icedream_weekly_xlsx(f):
            data = _parse_icedream_weekly_xlsx(f)
        else:
            data = parse_icedreams_file(f)
        month = data['month']
        if month in results:
            for product, vals in data['totals'].items():
                if product not in results[month]['totals']:
                    results[month]['totals'][product] = {'units': 0, 'value': 0, 'cartons': 0}
                for k in ['units', 'value', 'cartons']:
                    results[month]['totals'][product][k] += vals[k]
            for cust, pdata in data.get('by_customer', {}).items():
                if cust not in results[month]['by_customer']:
                    results[month]['by_customer'][cust] = {}
                for p, vals in pdata.items():
                    if p not in results[month]['by_customer'][cust]:
                        results[month]['by_customer'][cust][p] = {'units': 0, 'value': 0, 'cartons': 0}
                    for k in ['units', 'value', 'cartons']:
                        results[month]['by_customer'][cust][p][k] += vals.get(k, 0)
        else:
            results[month] = data

    # ── Supplement months with unattributed units from Format B files ────
    # Some weekly .xlsx files produce flat totals but empty by_customer
    # (e.g. icedream_mar_w10_11.xlsx lacks bold customer-summary rows).
    # The Format B file (sales_week_12.xls or sales_week_12.xlsx) covers the
    # same weeks with full per-customer data.  We take all week columns EXCEPT
    # the LAST one (the last week is already attributed via weekly xlsx).
    # Prefer .xlsx version (no xlrd needed); skip .xls if .xlsx exists.
    format_b_files = []
    for f in sorted(folder.glob('*.xlsx')):
        if f.name.startswith('~') or 'stock' in f.name.lower():
            continue
        if _is_format_b_multiweek_xlsx(f):
            format_b_files.append(f)
    # Add .xls files that don't have a .xlsx counterpart already listed
    xlsx_stems = {f.stem for f in format_b_files}
    for f in sorted(folder.glob('*.xls')):
        if f.name.startswith('~') or 'stock' in f.name.lower():
            continue
        if f.stem not in xlsx_stems:
            format_b_files.append(f)

    for f in sorted(format_b_files):
        if f.name.startswith('~') or 'stock' in f.name.lower():
            continue
        format_b = parse_format_b_xls(f)
        if not format_b:
            continue
        n_weeks = len(next(iter(format_b.values()), {}))
        if n_weeks < 2:
            continue  # nothing to supplement (no earlier weeks)
        weeks_to_supplement = range(n_weeks - 1)  # all except last

        for month_str, mdata in results.items():
            flat_total = sum(v.get('units', 0) for v in mdata.get('totals', {}).values())
            cust_total = sum(
                sum(p.get('units', 0) for p in prods.values())
                for prods in mdata.get('by_customer', {}).values()
            )
            if flat_total <= cust_total:
                continue  # already fully attributed — nothing to supplement

            # Merge early-week customer data from Format B into by_customer.
            # Format B returns English chain names; they pass through
            # extract_customer_name() unchanged and resolve via _CUSTOMER_EN_TO_CC_ID.
            for chain_en, weeks in format_b.items():
                for wk_i in weeks_to_supplement:
                    wk_data = weeks.get(wk_i, {})
                    for product, pd in wk_data.items():
                        units = pd.get('units', 0)
                        value = round(pd.get('value', 0.0), 2)
                        if units == 0:
                            continue
                        if chain_en not in mdata['by_customer']:
                            mdata['by_customer'][chain_en] = {}
                        if product not in mdata['by_customer'][chain_en]:
                            mdata['by_customer'][chain_en][product] = {'units': 0, 'value': 0.0, 'cartons': 0}
                        mdata['by_customer'][chain_en][product]['units'] += units
                        mdata['by_customer'][chain_en][product]['value'] += value

    return results


# ── Ma'ayan Price Table ───────────────────────────────────────────────────

# Maps Maayan raw chain names (שם רשת) → price DB customer name
_MAAYAN_CHAIN_TO_PRICEDB = {
    'דור אלון':                    'AMPM',
    'שוק פרטי':                    'שוק פרטי',
    'דלק מנטה':                    'דלק',
    'סונול':                       'סונול',
    'פז ילו':                      'פז יילו',
    'פז יילו':                     'פז יילו',
    'פז חברת נפט- סופר יודה':      'פז סופר יודה',
    'שפר את אלי לוי בע"מ':         'אלונית',
}

# Maps Hebrew product names in price DB → internal product keys
_PRICEDB_PRODUCT_MAP = {
    'גלידת חלבון וניל':            'vanilla',
    'גלידת חלבון מנגו':            'mango',
    'גלידת חלבון פיסטוק':          'pistachio',
    'גלידת חלבון שוקולד לוז':      'chocolate',
}

def _load_mayyan_price_table():
    """Load latest price DB file and return {product: {chain: price}} for Maayan.
    Falls back to B2B list price via pricing_engine for any missing product/chain combo.
    """
    price_dir = DATA_DIR / 'price data'
    if not price_dir.exists():
        return {}

    # Use most recently modified price db file
    candidates = sorted(
        [f for f in price_dir.glob('price db*.xlsx') if not f.name.startswith('~')],
        key=lambda f: f.stat().st_mtime,
        reverse=True,
    )
    if not candidates:
        return {}

    try:
        # Detect the header row: scan for the row that contains 'לקוח' and 'מחיר'
        _probe = pd.read_excel(candidates[0], header=None)
        header_row = 0
        for _i, _row in _probe.iterrows():
            _vals = [str(v) for v in _row.values]
            if any('לקוח' in v for v in _vals) and any('מחיר' in v for v in _vals):
                header_row = _i
                break
        df = pd.read_excel(candidates[0], header=header_row)
        price_col = next((c for c in df.columns if 'מחיר' in str(c) and 'מכירה' in str(c)), None)
        cust_col  = next((c for c in df.columns if 'לקוח' in str(c)), None)
        # 'מוצרים' is old column name; file uses 'שם פריט' or 'פריט' in newer exports
        prod_col  = next((c for c in df.columns if 'מוצרים' in str(c) or 'פריט' in str(c)), None)
        dist_col  = next((c for c in df.columns if 'מפיץ' in str(c)), None)
        if not all([price_col, cust_col, prod_col, dist_col]):
            return {}

        # Filter to Maayan rows only
        may_df = df[df[dist_col].astype(str).str.contains('מעיין', na=False)]

        # Build {product_key: {pricedb_customer: price}}
        table = {}
        for _, row in may_df.iterrows():
            prod_key = _PRICEDB_PRODUCT_MAP.get(str(row[prod_col]).strip())
            if not prod_key:
                continue
            cust = str(row[cust_col]).strip()
            price = float(row[price_col]) if row[price_col] else 0
            if prod_key not in table:
                table[prod_key] = {}
            table[prod_key][cust] = price
        return table
    except Exception:
        return {}


def _mayyan_chain_price(price_table, chain_raw, product):
    """Look up actual selling price for a Maayan chain + product.

    Three-tier lookup:
      1. master_data JSONB via get_customer_price (SSOT)
      2. Ma'ayan price-DB Excel file (legacy fallback)
      3. B2B list price via pricing_engine (last resort)
    """
    chain_str = str(chain_raw).strip()

    # 1. Try master_data JSONB — resolve Hebrew chain to English customer name
    customer_en = extract_customer_name(chain_str)
    if customer_en:
        from pricing_engine import _md_sale_prices, _load_md_pricing
        _load_md_pricing()
        # Check if MD has a price for this (sku, customer, distributor) combo
        for (s, c, d), price in _md_sale_prices.items():
            if s == product and c == customer_en:
                return price

    # 2. Fallback: Ma'ayan price-DB Excel file
    pricedb_cust = _MAAYAN_CHAIN_TO_PRICEDB.get(chain_str)
    if pricedb_cust and product in price_table:
        price = price_table[product].get(pricedb_cust)
        if price:
            return price

    # 3. Last resort: B2B list price
    return get_b2b_price_safe(product)


# ── Ma'ayan Parser ───────────────────────────────────────────────────────

def parse_mayyan_file(filepath, price_table=None):
    """Parse Ma'ayan detailed distribution report.

    Always reads from the detail sheet (דוח_הפצה_גלידות_טורבו__אל_פירוט).
    The file also contains a pivot/summary sheet (טבלת ציר) whose raw sum
    double-counts rows — we never use it.

    Sheet selection priority:
      1. Any sheet whose name contains 'פירוט'  (most specific — the detail sheet)
      2. Any sheet whose name contains 'דוח'    (secondary — also the detail sheet)
      3. Reject sheets whose name contains pivot/summary keywords (ציר, סיכום, summary)
      4. Last sheet as final fallback
    """
    import pandas as pd
    if price_table is None:
        price_table = _load_mayyan_price_table()
    wb_meta = load_workbook(filepath, read_only=True)
    sheet_names = wb_meta.sheetnames
    wb_meta.close()

    # Sheets to skip — pivot/summary tabs that aggregate rows
    SKIP_KEYWORDS = ('ציר', 'סיכום', 'summary', 'pivot', 'totals')

    # Priority 1: sheet name contains 'פירוט'
    detail_sheet = next(
        (s for s in sheet_names if 'פירוט' in s),
        None
    )
    # Priority 2: sheet name contains 'דוח' but is NOT a skip-sheet
    if not detail_sheet:
        detail_sheet = next(
            (s for s in sheet_names
             if 'דוח' in s and not any(kw in s.lower() for kw in SKIP_KEYWORDS)),
            None
        )
    # Priority 3: any non-pivot sheet
    if not detail_sheet:
        detail_sheet = next(
            (s for s in sheet_names
             if not any(kw in s.lower() for kw in SKIP_KEYWORDS)),
            None
        )
    # Final fallback: last sheet
    if not detail_sheet:
        detail_sheet = sheet_names[-1]

    import os
    print(f"  [Mayyan] {os.path.basename(filepath)} → sheet: '{detail_sheet}'")
    df = pd.read_excel(filepath, sheet_name=detail_sheet)

    month_col = next((c for c in df.columns if 'חודש' in str(c)), None)
    week_col = next((c for c in df.columns if 'שבועי' in str(c)), None)
    product_col = next((c for c in df.columns if 'פריט' in str(c)), None)
    units_col = next((c for c in df.columns if 'בודדים' in str(c)), None)
    chain_col = next((c for c in df.columns if 'רשת' in str(c)), None)
    type_col = next((c for c in df.columns if 'סוג' in str(c)), None)
    branch_col = next((c for c in df.columns if 'חשבון' in str(c) and 'שם' in str(c)), None)
    # Fallback columns used when branch_col collapses to the chain name
    # (happens in W14/W15 where Ma'ayan put 'שוק פרטי' in שם חשבון for every private-market row)
    address_col = next((c for c in df.columns if 'כתובת' in str(c)), None)
    city_col = next((c for c in df.columns if str(c).startswith('עיר')), None)
    acct_key_col = next((c for c in df.columns if 'מפתח' in str(c) and 'חשבון' in str(c)), None)

    if not all([product_col, units_col]):
        return {}
    if not month_col and not week_col:
        return {}

    df['product'] = df[product_col].apply(lambda x: _validated_product(classify_product(x)))
    df = df[df['product'].notna()]

    if month_col:
        # Standard monthly format — map Hebrew month names to standard keys
        from config import HEBREW_MONTH_NAMES, MONTH_ORDER
        month_map = {}
        for m in df[month_col].unique():
            ms = str(m)
            matched = False
            # Check Hebrew and English month names dynamically
            for heb_name, eng_name in HEBREW_MONTH_NAMES.items():
                if heb_name in ms or eng_name in ms:
                    # Find the matching full month key in MONTH_ORDER
                    for full_key in MONTH_ORDER:
                        if eng_name in full_key:
                            month_map[m] = full_key
                            matched = True
                            break
                    break
            if not matched:
                month_map[m] = str(m)
        df['month_std'] = df[month_col].map(month_map)
    else:
        # Weekly format — no month column; infer month from filename or sheet names
        from config import FILENAME_MONTH_KEYWORDS, WEEK_TO_MONTH
        fname = os.path.basename(str(filepath)).lower()
        inferred_month = None
        for kw, month_val in FILENAME_MONTH_KEYWORDS.items():
            if kw in fname:
                inferred_month = month_val
                break
        if not inferred_month:
            # Try sheet names
            for sn in sheet_names:
                for kw, month_val in FILENAME_MONTH_KEYWORDS.items():
                    if kw in sn.lower():
                        inferred_month = month_val
                        break
                if inferred_month:
                    break
        if not inferred_month and week_col:
            # Infer month from week numbers using central registry
            weeks_in_data = df[week_col].dropna().unique()
            for w in weeks_in_data:
                try:
                    w_int = int(w)
                    if w_int in WEEK_TO_MONTH:
                        inferred_month = WEEK_TO_MONTH[w_int]
                        break
                except (ValueError, TypeError):
                    pass
        if not inferred_month:
            inferred_month = 'Unknown'
        df['month_std'] = inferred_month
    results = {}

    for month in df['month_std'].unique():
        mdf = df[df['month_std'] == month]

        # Compute value per row using per-chain actual prices from price DB
        totals = {}
        if chain_col and price_table:
            for product in mdf['product'].unique():
                pdf = mdf[mdf['product'] == product]
                total_units = int(pdf[units_col].sum())
                total_value = 0.0
                for _, row in pdf.iterrows():
                    row_units = row[units_col] if row[units_col] else 0
                    chain_raw = row[chain_col] if chain_col else ''
                    unit_price = _mayyan_chain_price(price_table, chain_raw, product)
                    total_value += row_units * unit_price
                totals[product] = {
                    'units': total_units,
                    'value': round(total_value, 2),
                    'transactions': int(len(pdf))
                }
        else:
            for product in mdf['product'].unique():
                pdf = mdf[mdf['product'] == product]
                totals[product] = {
                    'units': int(pdf[units_col].sum()),
                    'value': 0,
                    'transactions': int(len(pdf))
                }

        by_chain = {}
        if chain_col:
            for _, row in mdf.groupby([chain_col, 'product']).agg(
                total_units=(units_col, 'sum')
            ).reset_index().iterrows():
                chain = str(row[chain_col])
                if chain not in by_chain:
                    by_chain[chain] = {}
                by_chain[chain][row['product']] = int(row['total_units'])

        by_type = {}
        if type_col:
            for _, row in mdf.groupby([type_col, 'product']).agg(
                total_units=(units_col, 'sum')
            ).reset_index().iterrows():
                ct = str(row[type_col])
                if ct not in by_type:
                    by_type[ct] = {}
                by_type[ct][row['product']] = int(row['total_units'])

        by_account = {}
        if branch_col:
            for _, row in mdf.iterrows():
                acct_raw = row.get(branch_col) if branch_col in row.index else None
                if not acct_raw:
                    continue
                acct = str(acct_raw).strip()
                chain_name = str(row[chain_col]).strip() if chain_col and chain_col in row.index else ''
                # Fallback: when branch_col collapses to the chain name (W14/W15 bug in
                # Ma'ayan's export — all שוק פרטי rows have שם חשבון='שוק פרטי'),
                # compose a per-branch name from address+city, disambiguated by acct key.
                if acct == chain_name:
                    addr = str(row.get(address_col) or '').strip() if address_col else ''
                    city = str(row.get(city_col) or '').strip() if city_col else ''
                    key_val = str(row.get(acct_key_col) or '').strip() if acct_key_col else ''
                    # Tiv Taam stores: when chain collapses to שוק פרטי,
                    # identify by acct_key prefix 31900xx
                    if chain_name == 'שוק פרטי' and key_val.startswith('31900'):
                        chain_name = 'טיב טעם'
                    label_parts = [p for p in (addr, city) if p and p.lower() != 'nan']
                    label = ', '.join(label_parts)
                    if label and key_val:
                        acct = f"{chain_name} — {label} (#{key_val})"
                    elif label:
                        acct = f"{chain_name} — {label}"
                    elif key_val:
                        acct = f"{chain_name} — #{key_val}"
                product = row.get('product')
                if not product:
                    continue
                row_units = int(row[units_col]) if row.get(units_col) else 0
                unit_price = _mayyan_chain_price(price_table, chain_name, product)
                key = (chain_name, acct)
                if key not in by_account:
                    by_account[key] = {}
                if product not in by_account[key]:
                    by_account[key][product] = {'units': 0, 'value': 0.0}
                by_account[key][product]['units'] += row_units
                by_account[key][product]['value'] = round(
                    by_account[key][product]['value'] + row_units * unit_price, 2
                )

        branches = set()
        if branch_col:
            for b in mdf[branch_col].dropna().unique():
                branches.add(str(b).strip())

        results[month] = {
            'totals': totals,
            'by_chain': by_chain,
            'by_account': by_account,
            'by_customer_type': by_type,
            'branches': branches,
        }

    return results


def parse_all_mayyan():
    folder = DATA_DIR / 'mayyan'
    if not folder.exists():
        return {}
    price_table = _load_mayyan_price_table()
    results = {}
    for f in sorted(folder.glob('*.xlsx')):
        if f.name.startswith('~'):
            continue
        # Skip stock files — handled separately by get_distributor_inventory()
        if 'stock' in f.name.lower():
            continue
        data = parse_mayyan_file(f, price_table=price_table)
        for month, mdata in data.items():
            if month not in results:
                results[month] = mdata
            else:
                # Merge: accumulate totals and by_account across multiple files for the same month
                existing = results[month]
                # Merge totals
                for sku, vals in mdata.get('totals', {}).items():
                    if sku not in existing.setdefault('totals', {}):
                        existing['totals'][sku] = {'units': 0, 'value': 0.0}
                    existing['totals'][sku]['units'] += vals.get('units', 0)
                    existing['totals'][sku]['value'] = round(
                        existing['totals'][sku]['value'] + vals.get('value', 0.0), 2)
                # Merge by_account
                for key, products in mdata.get('by_account', {}).items():
                    if key not in existing.setdefault('by_account', {}):
                        existing['by_account'][key] = {}
                    for sku, vals in products.items():
                        if sku not in existing['by_account'][key]:
                            existing['by_account'][key][sku] = {'units': 0, 'value': 0.0}
                        existing['by_account'][key][sku]['units'] += vals.get('units', 0)
                        existing['by_account'][key][sku]['value'] = round(
                            existing['by_account'][key][sku]['value'] + vals.get('value', 0.0), 2)
                # Merge by_chain
                for chain, products in mdata.get('by_chain', {}).items():
                    if chain not in existing.setdefault('by_chain', {}):
                        existing['by_chain'][chain] = {}
                    for sku, units in products.items():
                        existing['by_chain'][chain][sku] = existing['by_chain'][chain].get(sku, 0) + units
                # Merge by_customer_type
                for ctype, products in mdata.get('by_customer_type', {}).items():
                    if ctype not in existing.setdefault('by_customer_type', {}):
                        existing['by_customer_type'][ctype] = {}
                    for sku, units in products.items():
                        existing['by_customer_type'][ctype][sku] = existing['by_customer_type'][ctype].get(sku, 0) + units
                # Merge branches
                existing.setdefault('branches', set()).update(mdata.get('branches', set()))
    return results


# ── Production Parser ─────────────────────────────────────────────────

def get_production_data():
    """Returns known production data. Will be extended to parse files."""
    production = {
    }
    folder = DATA_DIR / 'production'
    if folder.exists():
        for f in sorted(folder.glob('*.xlsx')):
            pass  # TODO: parse production files when format is known
    return production


# ── Karfree Warehouse Inventory Parser ────────────────────────────────

def _classify_product_karfree(text):
    """Classify product from reversed Hebrew PDF text."""
    t = text.lower()
    # Exclude non-Raito products (דובאי = יאבוד reversed, באגסו = וסגאב reversed)
    if 'יאבוד' in t or 'וסגאב' in t:
        return None
    if 'וגנמ' in t or 'mango' in t:
        return 'mango'
    if 'דלוקוש' in t or 'chocolate' in t:
        return 'chocolate'
    if 'לינו' in t or 'vanilla' in t:
        return 'vanilla'
    if 'קוטסיפ' in t or 'וקטסיפ' in t or 'pistachio' in t:
        return 'pistachio'
    if 'תגדגמ' in t or 'magadat' in t:
        return 'magadat'
    if 'תגוע' in t or 'dream' in t or 'cake' in t:
        return 'dream_cake'
    return None


def parse_karfree_inventory():
    """Parse Karfree cold storage PDF inventory reports."""
    folder = DATA_DIR / 'karfree'
    if not folder.exists():
        return {}

    # Glob *.pdf plus non-pdf files that are actually PDFs (missing/wrong extension)
    pdf_files = list(folder.glob('*.pdf'))
    for f in folder.iterdir():
        if f.is_file() and f.suffix != '.pdf' and f not in pdf_files:
            try:
                with open(f, 'rb') as fh:
                    if fh.read(5) == b'%PDF-':
                        pdf_files.append(f)
            except Exception:
                pass
    if not pdf_files:
        return {}

    try:
        import pdfplumber
    except ImportError:
        print("  Warning: pdfplumber not installed, skipping inventory")
        return {}

    # Pick the file with the latest report date (not mtime — mtime is unreliable).
    # Parse all candidates, return the one with the newest date.
    best_result = None
    best_date = (0, 0, 0)

    for filepath in pdf_files:
        results = {
            'report_date': None,
            'products': {},
            'total_units': 0,
            'total_pallets': 0,
        }

        try:
            with pdfplumber.open(filepath) as pdf:
                full_text = ''
                for page in pdf.pages:
                    full_text += page.extract_text() + '\n'
        except Exception:
            continue

        lines = full_text.split('\n')
        current_product = None

        for line in lines:
            if ':ךיראתל ןוכנ' in line:
                date_match = re.search(r'(\d{2}/\d{2}/\d{4})', line)
                if date_match:
                    results['report_date'] = date_match.group(1)

            if 'טירפ' in line and ':' in line:
                product = _classify_product_karfree(line)
                current_product = product  # Reset even if None (e.g. Dubai products)
                if product and product not in results['products']:
                    results['products'][product] = {
                        'units': 0, 'pallets': 0, 'batches': []
                    }

            if current_product and re.match(r'^\s*0\s+000017', line):
                parts = line.split()
                try:
                    packages = None
                    for i, p in enumerate(parts):
                        if p == '0.00' and i + 1 < len(parts):
                            packages = int(parts[i + 1])
                            break
                    if packages:
                        # Vanilla (לינו) has both 6-pack and 10-pack SKUs:
                        # non-240 pallets are old 6-packs, 240 pallets are 10-packs
                        if current_product == 'vanilla':
                            factor = 10 if packages == 240 else 6
                        else:
                            factor = 10  # all other ice cream products are 10 units/carton
                        actual_units = packages * factor
                        dates = re.findall(r'\d{2}/\d{2}/\d{4}', line)
                        batch = {
                            'packages': packages,
                            'factor': factor,
                            'units': actual_units,
                            'expiry': dates[0] if len(dates) > 0 else None,
                            'production': dates[1] if len(dates) > 1 else None,
                            'entry': dates[2] if len(dates) > 2 else None,
                        }
                        results['products'][current_product]['batches'].append(batch)
                except (ValueError, IndexError):
                    pass

            if 'טירפל' in line and 'כ"הס' in line:
                total_match = re.search(r'(\d[\d\s,]*\d)\s+(\d+)\s+:טירפל', line)
                if total_match and current_product:
                    pallets_count = int(total_match.group(2))
                    results['products'][current_product]['pallets'] = pallets_count

        for p, pdata in results['products'].items():
            pdata['units'] = sum(b['units'] for b in pdata['batches'])
            pdata['packages'] = sum(b['packages'] for b in pdata['batches'])
            results['total_units'] += pdata['units']
            results['total_pallets'] += pdata['pallets']

        # Compare by report date
        if results.get('report_date') and results.get('total_units', 0) > 0:
            date_key = _stock_report_date_key(results['report_date'])
            if date_key > best_date:
                best_date = date_key
                best_result = results

    return best_result or {}


def get_warehouse_data():
    """Returns warehouse inventory data from Karfree reports."""
    return parse_karfree_inventory()


# ── Distributor Inventory Parser ──────────────────────────────────────

def parse_distributor_stock(filepath):
    """Parse a distributor stock Excel file (Icedream / Ma'ayan format)."""
    from pathlib import Path
    fp = Path(filepath)
    # Handle files without .xlsx extension
    if fp.suffix not in ('.xlsx', '.xlsm', '.xltx', '.xltm'):
        import shutil, tempfile
        tmp = Path(tempfile.mktemp(suffix='.xlsx'))
        shutil.copy2(fp, tmp)
        filepath = tmp
    wb = load_workbook(filepath)
    ws = wb[wb.sheetnames[0]]
    results = {'products': {}, 'total_units': 0, 'report_date': None}

    for row_idx in range(1, min(5, ws.max_row + 1)):
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                date_match = re.search(r'(\d{1,2}/\d{1,2}/\d{4})', str(val))
                if date_match:
                    results['report_date'] = date_match.group(1)
                    break
        if results['report_date']:
            break

    name_col = None
    qty_col = None
    header_row = None
    for row_idx in range(1, min(10, ws.max_row + 1)):
        for col_idx in range(1, ws.max_column + 1):
            val = str(ws.cell(row=row_idx, column=col_idx).value or '')
            if 'שם פריט' in val or ('פריט' in val and 'מפתח' not in val and 'בר' not in val):
                name_col = col_idx
                header_row = row_idx
            if 'מלאי' in val or 'כמות' in val or 'יתרת' in val:
                qty_col = col_idx
        if name_col and qty_col:
            break

    if not name_col or not qty_col:
        wb.close()
        return results

    is_units_format = False
    start_row = (header_row + 1) if header_row else 5
    for test_row in range(start_row, min(start_row + 3, ws.max_row + 1)):
        test_name = str(ws.cell(row=test_row, column=name_col).value or '')
        if re.search(r'\d+/\d+', test_name) and not re.search(r'\*\s*\d+\s*יח', test_name):
            is_units_format = True
            break

    for row_idx in range(start_row, ws.max_row + 1):
        item_name = ws.cell(row=row_idx, column=name_col).value
        qty_val = ws.cell(row=row_idx, column=qty_col).value

        if not item_name or qty_val is None:
            continue
        item_name = str(item_name)

        if 'סה"כ' in item_name:
            continue

        product = _validated_product(classify_product(item_name))
        if not product:
            continue

        qty = float(qty_val)
        if is_units_format:
            units = max(0, int(qty))
            factor = 1
            cartons = qty
        else:
            factor = extract_units_per_carton(item_name)
            cartons = qty
            units = math.ceil(cartons * factor)

        if units <= 0:
            continue

        if product not in results['products']:
            results['products'][product] = {'units': 0, 'cartons': 0, 'factor': factor}
        results['products'][product]['units'] += units
        results['products'][product]['cartons'] += cartons
        results['total_units'] += units

    wb.close()
    return results


def _stock_report_date_key(date_str):
    """Convert 'dd/mm/yyyy' report date to sortable (yyyy, mm, dd) tuple."""
    if not date_str:
        return (0, 0, 0)
    try:
        parts = date_str.split('/')
        return (int(parts[2]), int(parts[1]), int(parts[0]))
    except (IndexError, ValueError):
        return (0, 0, 0)


def get_distributor_inventory():
    """Parse distributor stock files from icedream and mayyan folders.

    Picks the stock file with the latest report_date (not file mtime)
    for each distributor.
    """
    dist_inv = {}

    def _find_best_stock(folders, dist_key):
        """Parse all stock files in folders, return the one with latest report date."""
        best = None
        best_date = (0, 0, 0)
        for folder in folders:
            if not folder.exists():
                continue
            # Case-insensitive stock file matching
            stock_files = [f for f in folder.iterdir() if 'stock' in f.name.lower() and not f.is_dir()]
            for f in stock_files:
                if f.name.startswith('~'):
                    continue
                try:
                    data = parse_distributor_stock(f)
                    if data.get('products'):
                        date_key = _stock_report_date_key(data.get('report_date'))
                        if date_key > best_date:
                            best_date = date_key
                            best = data
                except Exception:
                    pass
        if best:
            dist_inv[dist_key] = best

    _find_best_stock([DATA_DIR / 'icedreams', OUTPUT_DIR / 'icedream'], 'icedream')
    _find_best_stock([DATA_DIR / 'mayyan', OUTPUT_DIR / 'Maayan'], 'mayyan')

    return dist_inv


# ── Biscotti Parser ────────────────────────────────────────────────────

# Biscotti Dream Cake fallback price — used only when MD has no per-customer price
BISCOTTI_PRICE_DREAM_CAKE = get_b2b_price_safe('dream_cake_2')


def _biscotti_customer_price(branch_he: str) -> float:
    """Look up per-customer Biscotti price for dream_cake_2.

    Resolves Hebrew branch name to English customer name, then checks
    master_data JSONB via get_customer_price. Falls back to B2B list price.
    """
    customer_en = extract_customer_name(branch_he)
    if customer_en:
        return get_customer_price('dream_cake_2', customer_en, 'Biscotti')
    return BISCOTTI_PRICE_DREAM_CAKE

def _biscotti_week_to_month(week_num):
    """Map ISO week number to month key using central registry."""
    from config import WEEK_TO_MONTH
    return WEEK_TO_MONTH.get(week_num, f'W{week_num}')


def _parse_biscotti_file(filepath):
    """Parse Biscotti sales report.

    **New format (primary)**: Single sheet 'סיכום כללי' with columns:
      - 'מספר לקוח' (internal ID — ignored)
      - 'שם לקוח'   (sale point name)
      - 'כמות'      (quantity)
      Skip row where שם לקוח == 'סה"כ כללי' (grand total).

    **Old format (fallback)**: Multi-sheet Excel with 'שבוע N' weekly sheets
      and/or a 'סיכום' summary sheet (branch × week totals).

    Week-to-month: extracted from filename (e.g. week13.xlsx → W13 → March 2026).
    Falls back to 'March 2026' if no week number found.

    Returns {month_key: {'totals': {product: {'units', 'value'}},
                         'by_customer': {branch: {product: {'units', 'value'}}}}}
    """
    import os, re
    xl = pd.ExcelFile(filepath)
    fname = os.path.basename(filepath)
    print(f"  [Biscotti] {fname} → sheets: {xl.sheet_names}")

    # --- Determine month from filename week number ---
    week_match = re.search(r'week\s*(\d+)', fname, re.IGNORECASE)
    if week_match:
        month_key = _biscotti_week_to_month(int(week_match.group(1)))
    else:
        month_key = 'Unknown'  # no hardcoded fallback — will be filtered downstream

    branch_totals = {}  # {sale_point_name: units}

    # --- Try NEW format first: סיכום כללי with שם לקוח + כמות columns ---
    # Also accept any sheet that has 'שם לקוח' and 'כמות' column headers (e.g. PRI*.tmp sheets)
    SUMMARY_SHEET = 'סיכום כללי'
    new_format_parsed = False
    summary_sheet = next((s for s in xl.sheet_names if s.strip() == SUMMARY_SHEET), None)

    # If no exact sheet name match, scan all sheets for the expected column headers
    sheets_to_try = [summary_sheet] if summary_sheet else xl.sheet_names

    for sheet_name in sheets_to_try:
        if new_format_parsed:
            break
        df = pd.read_excel(filepath, sheet_name=sheet_name)
        # Find columns by name — also scan first 10 rows for header row
        name_col = None
        qty_col = None
        for col in df.columns:
            col_str = str(col).strip()
            if 'שם לקוח' in col_str:
                name_col = col
            elif 'כמות' in col_str:
                qty_col = col
        # If headers not in first row, scan rows for header row and re-read
        if name_col is None or qty_col is None:
            for idx in range(min(10, len(df))):
                row_vals = [str(v).strip() for v in df.iloc[idx]]
                if any('שם לקוח' in v for v in row_vals) and any('כמות' in v for v in row_vals):
                    df = pd.read_excel(filepath, sheet_name=sheet_name, header=idx + 1)
                    for col in df.columns:
                        col_str = str(col).strip()
                        if 'שם לקוח' in col_str:
                            name_col = col
                        elif 'כמות' in col_str:
                            qty_col = col
                    break
        if name_col is not None and qty_col is not None:
            new_format_parsed = True
            for _, row in df.iterrows():
                name = str(row[name_col]).strip()
                qty = row[qty_col]
                if not name or name == 'nan' or name == 'סה"כ כללי':
                    continue
                if pd.isna(qty):
                    continue
                try:
                    branch_totals[name] = branch_totals.get(name, 0) + int(qty)
                except (ValueError, TypeError):
                    continue

    # --- Fallback: OLD format (שבוע / סיכום sheets) ---
    if not new_format_parsed:
        SUMMARY_KEYWORD = 'סיכום'
        WEEK_KEYWORD = 'שבוע'
        old_summary = next((s for s in xl.sheet_names if SUMMARY_KEYWORD in s), None)
        week_sheets = [s for s in xl.sheet_names if WEEK_KEYWORD in s]

        if old_summary:
            df = pd.read_excel(filepath, sheet_name=old_summary, header=None)
            total_col = df.shape[1] - 1
            for i in range(3, len(df) - 1):
                branch = str(df.iloc[i, 0]).strip()
                total = df.iloc[i, total_col]
                if branch and branch != 'nan' and total and str(total) != 'nan':
                    branch_totals[branch] = int(total)
        elif week_sheets:
            for sheet in week_sheets:
                df = pd.read_excel(filepath, sheet_name=sheet, header=None)
                total_col = df.shape[1] - 1
                for i in range(2, len(df) - 1):
                    branch = str(df.iloc[i, 0]).strip()
                    total = df.iloc[i, total_col]
                    if branch and branch != 'nan' and total and str(total) != 'nan':
                        branch_totals[branch] = branch_totals.get(branch, 0) + int(total)

    if not branch_totals:
        return {}

    total_units = sum(branch_totals.values())

    by_customer = {}
    total_value = 0.0
    for branch, units in branch_totals.items():
        price = _biscotti_customer_price(branch)
        branch_value = round(units * price, 2)
        total_value += branch_value
        by_customer[branch] = {
            'dream_cake_2': {
                'units': units,
                'value': branch_value,
            }
        }
    total_value = round(total_value, 2)

    return {
        month_key: {
            'totals': {
                'dream_cake_2': {
                    'units': total_units,
                    'value': total_value,
                }
            },
            'by_customer': by_customer,
        }
    }


def parse_all_biscotti():
    """Parse all Biscotti distributor sales reports from data/biscotti/."""
    folder = DATA_DIR / 'biscotti'
    if not folder.exists():
        return {}

    results = {}
    for f in sorted(folder.glob('*.xlsx')):
        if f.name.startswith('~') or f.name.startswith('.'):
            continue
        # Skip temp files (e.g. PRI8C5F.tmp.xlsx)
        if '.tmp.' in f.name.lower():
            continue
        # Skip item-setup / admin files — only parse weekly sales reports
        # Identified by: sheet name contains 'שבוע' or 'סיכום' (old/new format),
        # OR filename contains 'week' (new PRI*.tmp sheet format),
        # OR sheet contains 'שם לקוח' + 'כמות' column headers
        try:
            xl = pd.ExcelFile(f)
            sheets = xl.sheet_names
            is_sales_report = any(
                'שבוע' in s or 'סיכום' in s for s in sheets
            )
            if not is_sales_report:
                # Also accept files named weekN.xlsx
                import re as _re
                if _re.search(r'week\s*\d+', f.name, _re.IGNORECASE):
                    is_sales_report = True
            if not is_sales_report:
                continue
        except Exception:
            continue

        try:
            data = _parse_biscotti_file(f)
            for month, mdata in data.items():
                if month not in results:
                    results[month] = mdata
                else:
                    # Merge: add units/value per branch
                    for branch, prods in mdata['by_customer'].items():
                        if branch not in results[month]['by_customer']:
                            results[month]['by_customer'][branch] = prods
                        else:
                            for p, pdata in prods.items():
                                existing = results[month]['by_customer'][branch].get(p, {})
                                results[month]['by_customer'][branch][p] = {
                                    'units': existing.get('units', 0) + pdata['units'],
                                    'value': existing.get('value', 0) + pdata['value'],
                                }
                    # Recalculate totals from per-branch values (already per-customer priced)
                    for p in results[month]['totals']:
                        results[month]['totals'][p]['units'] = sum(
                            results[month]['by_customer'][b].get(p, {}).get('units', 0)
                            for b in results[month]['by_customer']
                        )
                        results[month]['totals'][p]['value'] = round(sum(
                            results[month]['by_customer'][b].get(p, {}).get('value', 0)
                            for b in results[month]['by_customer']
                        ), 2)
        except Exception as e:
            print(f"  [Biscotti] Error parsing {f.name}: {e}")

    return results


# ── Data Consolidation ──────────────────────────────────────────────────

def consolidate_data():
    """Merge all data sources into a unified dataset."""
    print("Processing Icedreams reports...")
    icedreams = parse_all_icedreams()
    print(f"  Found {len(icedreams)} months")

    print("Processing Ma'ayan reports...")
    mayyan = parse_all_mayyan()
    print(f"  Found {len(mayyan)} months")

    print("Processing Biscotti reports...")
    biscotti = parse_all_biscotti()
    print(f"  Found {len(biscotti)} months")

    production = get_production_data()
    warehouse = get_warehouse_data()
    if warehouse:
        print(f"Processing Karfree inventory...")
        print(f"  Report date: {warehouse.get('report_date', 'N/A')}, Total: {warehouse.get('total_units', 0):,} units")

    dist_inv = get_distributor_inventory()
    for dist_name, dist_data in dist_inv.items():
        print(f"Processing {dist_name} stock...")
        print(f"  Report date: {dist_data.get('report_date', 'N/A')}, Total: {dist_data.get('total_units', 0):,} units")

    all_months = sorted(
        [m for m in set(list(icedreams.keys()) + list(mayyan.keys()) + list(biscotti.keys()))
         if m in MONTH_ORDER],
        key=lambda x: MONTH_ORDER.get(x, 99)
    )
    products = ['chocolate', 'vanilla', 'mango', 'magadat', 'dream_cake', 'dream_cake_2', 'pistachio']

    consolidated = {
        'months': all_months,
        'products': products,
        'monthly_data': {},
        'production': production,
        'warehouse': warehouse,
        'dist_inv': dist_inv,
    }

    for month in all_months:
        # Filter out internal/promo accounts (e.g. Oogiplatset) from customer dicts
        _ice_cust_raw = icedreams.get(month, {}).get('by_customer', {})
        _ice_cust = {k: v for k, v in _ice_cust_raw.items() if k not in INTERNAL_ACCOUNTS}
        _bisc_cust_raw = biscotti.get(month, {}).get('by_customer', {})
        _bisc_cust = {k: v for k, v in _bisc_cust_raw.items() if k not in INTERNAL_ACCOUNTS}

        month_data = {
            'icedreams': icedreams.get(month, {}).get('totals', {}),
            'mayyan': mayyan.get(month, {}).get('totals', {}),
            'icedreams_customers': _ice_cust,
            'mayyan_chains': mayyan.get(month, {}).get('by_chain', {}),
            'mayyan_accounts': mayyan.get(month, {}).get('by_account', {}),
            'mayyan_branches': mayyan.get(month, {}).get('branches', set()),
            'mayyan_types': mayyan.get(month, {}).get('by_customer_type', {}),
            'biscotti': biscotti.get(month, {}).get('totals', {}),
            'biscotti_customers': _bisc_cust,
            'combined': {},
        }

        for p in products:
            ice_units = month_data['icedreams'].get(p, {}).get('units', 0)
            may_units = month_data['mayyan'].get(p, {}).get('units', 0)
            bisc_units = month_data['biscotti'].get(p, {}).get('units', 0)
            ice_value = month_data['icedreams'].get(p, {}).get('value', 0)
            # Use actual per-chain Maayan revenue when available; fall back to estimate
            _may_actual = month_data['mayyan'].get(p, {}).get('value', 0)
            may_value = _may_actual if _may_actual > 0 else round(may_units * get_b2b_price_safe(p), 2)
            bisc_value = month_data['biscotti'].get(p, {}).get('value', 0)
            total_value = round(ice_value + may_value + bisc_value, 2)
            prod_cost_per_unit = get_production_cost(p)
            total_units = ice_units + may_units + bisc_units
            total_prod_cost = round(total_units * prod_cost_per_unit, 2)
            gross_margin = round(total_value - total_prod_cost, 2) if p != 'magadat' else 0

            month_data['combined'][p] = {
                'units': total_units,
                'icedreams_units': ice_units,
                'mayyan_units': may_units,
                'biscotti_units': bisc_units,
                'icedreams_value': ice_value,
                'mayyan_value': may_value,
                'biscotti_value': bisc_value,
                'total_value': total_value,
                'production_cost': total_prod_cost,
                'gross_margin': gross_margin,
            }

        filtered_custs = {}
        for cust, pdata in month_data.get('icedreams_customers', {}).items():
            total_u = sum(v.get('units', 0) for v in pdata.values())
            # Allow through negative-unit customers (returns/credit notes) so CC can
            # account for them — previously total_u > 0 silently dropped returns.
            if total_u != 0:
                for p, vals in pdata.items():
                    vals['value'] = round(vals.get('value', 0), 2)
                filtered_custs[cust] = pdata
        month_data['icedreams_customers'] = filtered_custs

        mayyan_chains_revenue = {}
        for chain, pdata in month_data.get('mayyan_chains', {}).items():
            mayyan_chains_revenue[chain] = {}
            for p, units in pdata.items():
                mayyan_chains_revenue[chain][p] = {
                    'units': units,
                    'value': round(units * get_b2b_price_safe(p), 2),
                }
        month_data['mayyan_chains_revenue'] = mayyan_chains_revenue

        # mayyan_accounts now carries {product: {units, value}} — pass through directly
        # (pricing was applied at parse time using _mayyan_chain_price per row)
        month_data['mayyan_accounts_revenue'] = month_data.get('mayyan_accounts', {})

        # Filter Biscotti customers: same hygiene as Icedream (drop zero-unit branches, round values)
        filtered_bisc = {}
        for branch, pdata in month_data.get('biscotti_customers', {}).items():
            total_u = sum(v.get('units', 0) for v in pdata.values())
            if total_u != 0:
                for p, vals in pdata.items():
                    vals['value'] = round(vals.get('value', 0), 2)
                filtered_bisc[branch] = pdata
        month_data['biscotti_customers'] = filtered_bisc

        consolidated['monthly_data'][month] = month_data

    # ── Phase 2: Enrich with DB IDs ─────────────────────────────────────
    consolidated['id_maps'] = _enrich_with_ids(consolidated)

    return consolidated


def _enrich_with_ids(consolidated):
    """Resolve all customer/product/distributor names to DB IDs.

    Returns an id_maps dict:
      {
        'customers':    {english_name: customer_id, ...},
        'products':     {sku_key: product_id, ...},
        'brands':       {brand_key: brand_id, ...},
        'distributors': {dist_key: distributor_id, ...},
        'available':    True/False   ← whether DB resolver was used
      }

    If DB is unavailable, returns {'available': False} and everything
    continues to work via string-based matching (backward compatible).
    """
    resolver = _get_resolver()
    if not resolver:
        return {'available': False}

    customer_id_map = {}   # english_name → customer_id
    product_id_map = {}    # sku_key → product_id
    brand_id_map = {}      # brand_key → brand_id
    dist_id_map = {}       # dist_key → distributor_id

    # ── Resolve products (from the known SKU list) ──
    for sku in consolidated.get('products', []):
        result = resolver.resolve_product_by_sku(sku)
        if result:
            product_id_map[sku] = result[0]  # product_id

    # ── Resolve distributors ──
    for dist_key in ('icedream', 'mayyan_froz', 'biscotti'):
        did = resolver.resolve_distributor(dist_key)
        if did:
            dist_id_map[dist_key] = did

    # ── Resolve brands ──
    for brand_key in ('turbo', 'danis'):
        bid = resolver.resolve_brand(brand_key)
        if bid:
            brand_id_map[brand_key] = bid

    # ── Resolve customer names (collect from all months) ──
    all_customer_names = set()
    for month, mdata in consolidated.get('monthly_data', {}).items():
        # Icedream customers (English names from extract_customer_name)
        all_customer_names.update(mdata.get('icedreams_customers', {}).keys())
        # Biscotti customers (Hebrew branch names — resolve to parent customer)
        all_customer_names.update(mdata.get('biscotti_customers', {}).keys())
        # Ma'ayan chains (Hebrew chain names)
        all_customer_names.update(mdata.get('mayyan_chains', {}).keys())

    for name in all_customer_names:
        if name in customer_id_map:
            continue
        result = resolver.resolve_customer(name)
        if result:
            customer_id_map[name] = result[0]  # customer_id

    # ── Report coverage ──
    n_cust = len(customer_id_map)
    n_cust_total = len(all_customer_names)
    n_prod = len(product_id_map)
    n_prod_total = len(consolidated.get('products', []))
    print(f"  ID enrichment: customers {n_cust}/{n_cust_total}, "
          f"products {n_prod}/{n_prod_total}, "
          f"distributors {len(dist_id_map)}, brands {len(brand_id_map)}")

    if resolver.unresolved_customers:
        print(f"  ⚠ {len(resolver.unresolved_customers)} unresolved customer names:")
        for name, count in sorted(resolver.unresolved_customers.items(), key=lambda x: -x[1])[:10]:
            print(f"    [{count}x] '{name}'")

    if resolver.unresolved_products:
        print(f"  ⚠ {len(resolver.unresolved_products)} unresolved product names:")
        for name, count in sorted(resolver.unresolved_products.items(), key=lambda x: -x[1])[:10]:
            print(f"    [{count}x] '{name}'")

    return {
        'available': True,
        'customers': customer_id_map,
        'products': product_id_map,
        'brands': brand_id_map,
        'distributors': dist_id_map,
    }
