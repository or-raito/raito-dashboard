#!/usr/bin/env python3
"""
RAITO — Phase 4: Cloud-Ready Dashboard Server (db_dashboard.py)

Flask application that generates the unified dashboard from PostgreSQL
and serves it as a web page. Designed for Gunicorn + Cloud Run.

Usage (local):
    cd scripts && python3 db/db_dashboard.py       # Flask dev server on :8080
    gunicorn --bind :8080 scripts.db.db_dashboard:app  # Production

Usage (Docker):
    docker build -t raito .
    docker run -p 8080:8080 -e DATABASE_URL=... raito

The `app` object is defined at module level so Gunicorn can find it via:
    scripts.db.db_dashboard:app
"""

import os
import sys
import logging
from pathlib import Path

# ── Path setup ────────────────────────────────────────────────────────────
# Ensure scripts/ is importable whether we're run from project root or scripts/
SCRIPTS_DIR = Path(__file__).resolve().parent.parent
PROJECT_ROOT = SCRIPTS_DIR.parent
sys.path.insert(0, str(SCRIPTS_DIR))
sys.path.insert(0, str(PROJECT_ROOT))

from flask import Flask, Response, send_file, request, jsonify, session

# Configure logging for Cloud Run (structured stdout)
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
)
log = logging.getLogger(__name__)

# ── Flask app (global — Gunicorn entry point) ─────────────────────────────
app = Flask(__name__)

# ── Auth (Phase 2) ───────────────────────────────────────────────────────
try:
    from db.auth import setup_auth, require_admin
    setup_auth(app)
    log.info("Auth module loaded — write routes require admin session")
except ImportError:
    log.warning("auth module not found — write routes UNPROTECTED")
    def require_admin(f):  # noqa: F811 — fallback no-op
        return f

# ── Geo Layer Blueprint ───────────────────────────────────────────────────
# Registers /api/geo/municipalities, /api/geo/choropleth, /api/geo/pos
try:
    from geo_api import geo_blueprint
    app.register_blueprint(geo_blueprint)
    log.info("Geo API blueprint registered (/api/geo/*)")
except Exception as _geo_err:
    log.warning(f"Geo API blueprint not loaded (run migration first): {_geo_err}")


def _generate_dashboard_html():
    """Parse Excel data files and run all dashboard generators, return HTML string.

    Uses parsers.consolidate_data() (Excel-based SSOT) — same pipeline as the
    local build. Does NOT require sales_transactions or any Phase-4 SQL tables.
    Uploaded files copied to data/ subfolders via _copy_to_data_folder() are
    automatically picked up on the next call to this function.
    """
    from master_data_parser import parse_master_data
    from unified_dashboard import (
        generate_unified_dashboard,
        DASHBOARD_PASSWORD,
        _js_hash,
    )
    from parsers import consolidate_data

    # 1. Parse all distributor Excel files from the data/ folder
    log.info("Parsing distributor Excel files...")
    data = consolidate_data()
    months = data.get('months', [])
    log.info(f"  Months loaded: {months}")

    for m in months:
        md = data['monthly_data'][m]
        combined = md.get('combined', {})
        total_u = sum(v.get('units', 0) for v in combined.values())
        total_r = sum(v.get('total_value', 0) for v in combined.values())
        log.info(f"  {m}: {total_u:,} units, ₪{total_r:,.0f} revenue")

    # 2. Parse master data (local files inside container)
    log.info("Parsing master data...")
    master_data = parse_master_data()

    # 3. Generate the unified dashboard HTML
    log.info("Generating unified dashboard HTML...")
    html = generate_unified_dashboard(data, master_data)

    # 4. Inject password hash
    pwd_hash = _js_hash(DASHBOARD_PASSWORD)
    html = html.replace('%%PWD_HASH%%', pwd_hash)

    log.info(f"Dashboard generated: {len(html) / 1024:.1f} KB")
    return html


# ── Cache ─────────────────────────────────────────────────────────────────
# In production the dashboard data changes at most weekly (new Excel ingestion).
# We cache the generated HTML and refresh on demand via /refresh.
# Phase 4: writes invalidate the cache and trigger a background rebuild so
# the next GET / is fast and other tabs reflect the changes.
import threading

_cached_html = None
_cache_lock = threading.Lock()
_cache_rebuilding = False


def _invalidate_cache():
    """Mark the cache as stale and start a background rebuild.

    Safe to call from any write route. The background thread pre-generates
    the HTML so the next GET / does not have to wait 10-30s.

    IMPORTANT: We do NOT null _cached_html here. The old (stale) HTML continues
    to be served while the rebuild runs. This prevents GET / from blocking
    synchronously when the rebuild takes a long time (e.g. dev env with no
    Excel data).
    """
    global _cached_html, _cache_rebuilding

    with _cache_lock:
        # Keep _cached_html as-is (serve stale while rebuilding)
        if _cache_rebuilding:
            log.info("Cache already rebuilding — skipping duplicate spawn")
            return
        _cache_rebuilding = True

    log.info("Cache invalidated — spawning background rebuild (non-blocking)")

    def _rebuild():
        global _cached_html, _cache_rebuilding
        try:
            log.info("Background rebuild starting...")
            html = _generate_dashboard_html()
            with _cache_lock:
                _cached_html = html
                _cache_rebuilding = False
            log.info("Background rebuild complete (%d KB)", len(html) // 1024)
        except Exception as exc:
            log.error("Background rebuild failed: %s", exc, exc_info=True)
            with _cache_lock:
                _cache_rebuilding = False

    t = threading.Thread(target=_rebuild, daemon=True)
    t.start()
    # Safety: reset the flag after 45s in case the thread hangs
    def _safety_reset():
        import time
        time.sleep(45)
        with _cache_lock:
            if _cache_rebuilding:
                log.warning("Background rebuild timed out after 45s — resetting flag")
                _cache_rebuilding = False  # noqa: F841
    threading.Thread(target=_safety_reset, daemon=True).start()


@app.route('/')
def index():
    """Serve the unified dashboard. Generates on first request, then cached."""
    global _cached_html
    if _cached_html is None:
        log.info("First request — generating dashboard...")
        _cached_html = _generate_dashboard_html()
    return Response(_cached_html, mimetype='text/html; charset=utf-8')


@app.route('/refresh')
def refresh():
    """Force-regenerate the dashboard from Excel files."""
    global _cached_html
    log.info("Refresh requested — regenerating dashboard...")
    _cached_html = _generate_dashboard_html()
    return Response(
        '<html><body><h2>Dashboard refreshed.</h2>'
        '<p><a href="/">Go to dashboard</a></p></body></html>',
        mimetype='text/html; charset=utf-8',
    )


@app.route('/api/data-check')
def data_check():
    """Debug endpoint — returns monthly totals from parsers.consolidate_data()
    without touching the HTML cache. Use to verify server-side data is correct
    before doing a full browser reload."""
    from parsers import consolidate_data
    data = consolidate_data()
    months = data.get('months', [])
    result = {}
    for m in months:
        md = data['monthly_data'][m]
        combined = md.get('combined', {})
        total_u = sum(v.get('units', 0) for v in combined.values())
        total_r = sum(v.get('total_value', 0) for v in combined.values())
        result[m] = {'units': total_u, 'revenue': round(total_r, 2)}
    return jsonify({'months': result, 'data_dir': str(
        __import__('config').DATA_DIR)})


@app.route('/health')
def health():
    """Health check for Cloud Run / load balancers."""
    return {'status': 'ok'}


# ── Weekly Override Helpers ───────────────────────────────────────────────────

_WEEK_LABELS = ["28/12","4/1","11/1","18/1","25/1","1/2","8/2","15/2","22/2","1/3","8/3","15/3","22/3"]


def _extract_week_overrides(distributor, filepath):
    """Extract weekly totals from an uploaded distributor file.

    Returns a LIST of dicts (one per week found in the file):
        [{'week_num': int, 'units': int, 'revenue': float, 'label': str}, ...]
    Returns [] if nothing could be extracted.

    Icedream: single-week file — reads D1 for week num, sums Raito product rows.
    Ma'ayan:  multi-week DETAIL sheet (דוח_הפצה_גלידות_טורבו__אל_פירוט) — iterates
              rows grouped by week column, prices each row via _mayyan_chain_price()
              from the price DB. NEVER reads from טבלת ציר (pivot — double-counts).
    Biscotti: multi-week — reads שבוע N (date) sheet names, maps start date to
              Raito week number via _WEEK_LABELS, prices at BISCOTTI_PRICE_DREAM_CAKE.
    """
    import re
    from openpyxl import load_workbook

    if distributor not in ('icedream', 'mayyan', 'biscotti'):
        return []

    results = []

    try:
        if distributor == 'icedream':
            wb = load_workbook(filepath)
            ws = wb.active

            # D1 contains e.g. 'שבוע 13 2026'
            d1_val = str(ws['D1'].value or '').strip()
            week_num = None
            if 'שבוע' in d1_val:
                m = re.search(r'שבוע\s+(\d+)', d1_val)
                if m:
                    week_num = int(m.group(1))

            if week_num and 1 <= week_num <= 52:
                total_units = 0
                total_revenue = 0.0
                for row_idx in range(3, ws.max_row + 1):
                    item_name = ws.cell(row_idx, 2).value
                    qty       = ws.cell(row_idx, 3).value
                    rev       = ws.cell(row_idx, 4).value
                    if not item_name:
                        continue
                    item_str = str(item_name).strip().lstrip("'")
                    if 'סה"כ' in item_str or not item_str:
                        continue
                    if 'טורבו' not in item_str and 'דרים קייק' not in item_str:
                        continue
                    if qty is not None:
                        try:
                            q = float(qty)
                            if q != 0:
                                from parsers import extract_units_per_carton
                                upc = extract_units_per_carton(item_str) or 1
                                total_units += int(round(-q * upc))
                        except (ValueError, TypeError):
                            pass
                    if rev is not None:
                        try:
                            r = float(rev)
                            if r != 0:
                                total_revenue += -r
                        except (ValueError, TypeError):
                            pass
                wb.close()
                if total_units > 0:
                    label = _WEEK_LABELS[week_num - 1] if 1 <= week_num <= len(_WEEK_LABELS) else f"W{week_num}"
                    results.append({'week_num': week_num, 'units': int(total_units),
                                    'revenue': round(total_revenue, 2), 'label': label})

        elif distributor == 'mayyan':
            # Reuses parsers.py logic — same sheet selection, column detection,
            # and per-row pricing via _mayyan_chain_price() from price DB.
            # NEVER reads from טבלת ציר (pivot sheet — double-counts rows).
            import sys as _sys
            import pandas as _pd
            _scripts_dir = str(Path(__file__).resolve().parent.parent)
            if _scripts_dir not in _sys.path:
                _sys.path.insert(0, _scripts_dir)
            from parsers import (
                _load_mayyan_price_table, _mayyan_chain_price,
                classify_product, _validated_product,
            )

            # Sheet selection — identical to parse_mayyan_file() in parsers.py
            _wb_meta = load_workbook(filepath, read_only=True)
            _sheet_names = _wb_meta.sheetnames
            _wb_meta.close()
            SKIP_KW = ('ציר', 'סיכום', 'summary', 'pivot', 'totals')
            _detail = next((s for s in _sheet_names if 'פירוט' in s), None)
            if not _detail:
                _detail = next((s for s in _sheet_names
                                if 'דוח' in s and not any(kw in s.lower() for kw in SKIP_KW)), None)
            if not _detail:
                _detail = next((s for s in _sheet_names
                                if not any(kw in s.lower() for kw in SKIP_KW)), None)
            if not _detail:
                _detail = _sheet_names[-1]

            df = _pd.read_excel(filepath, sheet_name=_detail)

            # Column detection — identical to parse_mayyan_file()
            week_col    = next((c for c in df.columns if 'שבועי'  in str(c)), None)
            product_col = next((c for c in df.columns if 'פריט'   in str(c)), None)
            units_col   = next((c for c in df.columns if 'בודדים' in str(c)), None)
            chain_col   = next((c for c in df.columns if 'רשת'    in str(c)), None)

            if all([week_col, product_col, units_col]):
                df['product'] = df[product_col].apply(
                    lambda x: _validated_product(classify_product(x))
                )
                df = df[df['product'].notna()]
                price_table = _load_mayyan_price_table()

                # Group by week number, calculate revenue per row from price DB
                for wn_raw in sorted(df[week_col].dropna().unique()):
                    try:
                        wn = int(wn_raw)
                    except (ValueError, TypeError):
                        continue
                    if not (1 <= wn <= 52):
                        continue
                    wdf = df[df[week_col] == wn_raw]
                    total_units = int(wdf[units_col].sum())
                    if total_units <= 0:
                        continue
                    total_revenue = 0.0
                    for _, row in wdf.iterrows():
                        row_units = row[units_col] if row.get(units_col) else 0
                        chain_raw = row[chain_col] if chain_col else ''
                        product   = row.get('product')
                        if product and row_units:
                            unit_price = _mayyan_chain_price(price_table, chain_raw, product)
                            total_revenue += row_units * unit_price
                    label = _WEEK_LABELS[wn - 1] if 1 <= wn <= len(_WEEK_LABELS) else f"W{wn}"
                    results.append({'week_num': wn, 'units': total_units,
                                    'revenue': round(total_revenue, 2), 'label': label})

        elif distributor == 'biscotti':
            # Biscotti format: multi-sheet Excel with 'שבוע N (DD.M–DD.M)' sheets.
            # Reuses BISCOTTI_PRICE_DREAM_CAKE from parsers.py (SSOT for pricing).
            # Week number is derived from the start date in each sheet name, mapped
            # to the Raito week number via _WEEK_LABELS boundaries.
            import sys as _sys
            import pandas as _pd
            from datetime import date as _date
            _scripts_dir = str(Path(__file__).resolve().parent.parent)
            if _scripts_dir not in _sys.path:
                _sys.path.insert(0, _scripts_dir)
            from parsers import BISCOTTI_PRICE_DREAM_CAKE

            # Build date → Raito week_num lookup from _WEEK_LABELS
            # Labels like "28/12", "4/1", ..., "22/3" = start-of-week dates (Dec 2025–Mar 2026)
            _boundaries = []
            for _wn, _lbl in enumerate(_WEEK_LABELS, start=1):
                _d, _mo = _lbl.split('/')
                _yr = 2025 if int(_mo) == 12 else 2026
                _boundaries.append((_wn, _date(_yr, int(_mo), int(_d))))

            def _date_to_raito_wn(d):
                """Return the Raito week number whose start date is ≤ d."""
                wn = None
                for week_num, boundary in _boundaries:
                    if d >= boundary:
                        wn = week_num
                return wn

            # Iterate שבוע N sheets (skip summary sheet)
            _wb_meta = load_workbook(filepath, read_only=True)
            _week_sheets = [s for s in _wb_meta.sheetnames if 'שבוע' in s]
            _wb_meta.close()

            for _sheet in _week_sheets:
                # Extract start date from sheet name: "שבוע 1 (18.3–20.3)" → 18.3
                _dm = re.search(r'\((\d+)\.(\d+)', _sheet)
                if not _dm:
                    continue
                _start_day, _start_month = int(_dm.group(1)), int(_dm.group(2))
                _start_yr = 2025 if _start_month == 12 else 2026
                _raito_wn = _date_to_raito_wn(_date(_start_yr, _start_month, _start_day))
                if not _raito_wn:
                    continue

                # Sum daily unit columns (col 1 to second-to-last; skip branch name + total col)
                df = _pd.read_excel(filepath, sheet_name=_sheet, header=None)
                _total = 0
                for _i in range(2, len(df) - 1):  # skip 2 header rows + last grand total row
                    _branch = str(df.iloc[_i, 0]).strip()
                    if not _branch or _branch == 'nan' or 'סה"כ' in _branch:
                        continue
                    for _val in df.iloc[_i, 1:-1]:  # skip first (branch) and last (formula total)
                        try:
                            _v = float(_val)
                            if not _pd.isna(_v) and _v > 0:
                                _total += int(_v)
                        except (ValueError, TypeError):
                            pass

                if _total > 0:
                    _label = _WEEK_LABELS[_raito_wn - 1] if 1 <= _raito_wn <= len(_WEEK_LABELS) else f"W{_raito_wn}"
                    _revenue = round(_total * BISCOTTI_PRICE_DREAM_CAKE, 2)
                    results.append({'week_num': _raito_wn, 'units': _total,
                                    'revenue': _revenue, 'label': _label})

    except Exception as e:
        log.warning(f"Failed to extract week overrides for {distributor}: {e}")

    return results


def _extract_week_override(distributor, filepath):
    """Backward-compatible single-result wrapper. Returns the highest week_num result or None."""
    results = _extract_week_overrides(distributor, filepath)
    if not results:
        return None
    # Return the highest week number extracted (most recent week)
    return max(results, key=lambda r: r['week_num'])


# ── Upload page ───────────────────────────────────────────────────────────────

_UPLOAD_HTML = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>RAITO — Data Upload</title>
<style>
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body {
    font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
    background: #0f1117; color: #e2e8f0; min-height: 100vh;
    display: flex; align-items: center; justify-content: center; padding: 24px;
  }
  .card {
    background: #1a1d27; border: 1px solid #2d3148; border-radius: 16px;
    width: 100%; max-width: 560px; padding: 40px; box-shadow: 0 20px 60px rgba(0,0,0,0.5);
  }
  .logo { font-size: 13px; letter-spacing: 3px; color: #6c7aaa; text-transform: uppercase; margin-bottom: 8px; }
  h1 { font-size: 26px; font-weight: 700; color: #fff; margin-bottom: 6px; }
  .subtitle { font-size: 14px; color: #6c7aaa; margin-bottom: 32px; }
  .drop-zone {
    border: 2px dashed #2d3148; border-radius: 12px; padding: 40px 24px;
    text-align: center; cursor: pointer; transition: all 0.2s;
    background: #12151f;
  }
  .drop-zone:hover, .drop-zone.drag-over {
    border-color: #5b6af0; background: #1a1e30;
  }
  .drop-icon { font-size: 36px; margin-bottom: 12px; }
  .drop-text { font-size: 15px; color: #a0aec0; }
  .drop-text span { color: #5b6af0; cursor: pointer; text-decoration: underline; }
  .file-name {
    display: none; margin-top: 12px; font-size: 13px; color: #68d391;
    background: #0d1f17; border: 1px solid #1a4731; border-radius: 8px; padding: 8px 12px;
  }
  .options { margin-top: 24px; display: flex; flex-direction: column; gap: 16px; }
  .field label { font-size: 12px; color: #6c7aaa; letter-spacing: 0.5px; text-transform: uppercase; margin-bottom: 6px; display: block; }
  select {
    width: 100%; background: #12151f; border: 1px solid #2d3148; border-radius: 8px;
    color: #e2e8f0; padding: 10px 14px; font-size: 14px; outline: none;
    appearance: none;
  }
  select:focus { border-color: #5b6af0; }
  .toggle-row { display: flex; align-items: center; gap: 10px; cursor: pointer; }
  .toggle-row input[type=checkbox] { width: 16px; height: 16px; accent-color: #5b6af0; cursor: pointer; }
  .toggle-label { font-size: 14px; color: #a0aec0; }
  .toggle-label small { display: block; font-size: 12px; color: #4a5568; margin-top: 2px; }
  .btn {
    width: 100%; padding: 14px; background: #5b6af0; color: #fff; border: none;
    border-radius: 10px; font-size: 15px; font-weight: 600; cursor: pointer;
    margin-top: 28px; transition: background 0.2s; letter-spacing: 0.3px;
  }
  .btn:hover { background: #4a58d8; }
  .btn:disabled { background: #2d3148; color: #4a5568; cursor: not-allowed; }
  .result {
    display: none; margin-top: 24px; border-radius: 12px; padding: 20px;
    font-size: 13px; line-height: 1.7;
  }
  .result.success { background: #0d1f17; border: 1px solid #1a4731; color: #68d391; }
  .result.error   { background: #1f0d0d; border: 1px solid #4a1a1a; color: #fc8181; }
  .result.skipped { background: #1a1a0d; border: 1px solid #4a4a1a; color: #f6e05e; }
  .result-title { font-weight: 700; font-size: 15px; margin-bottom: 8px; }
  .result table { width: 100%; border-collapse: collapse; margin-top: 8px; }
  .result td { padding: 3px 0; }
  .result td:first-child { color: #718096; padding-right: 16px; }
  .spinner { display: none; text-align: center; margin-top: 16px; color: #6c7aaa; font-size: 14px; }
  .back-link { display: block; text-align: center; margin-top: 24px; font-size: 13px; color: #5b6af0; text-decoration: none; }
  .back-link:hover { text-decoration: underline; }
</style>
</head>
<body>
<div class="card">
  <div class="logo">Raito</div>
  <h1>Upload Distributor File</h1>
  <p class="subtitle">Drop a sales or stock file — it loads directly into the database.</p>

  <div class="drop-zone" id="dropZone">
    <div class="drop-icon">📂</div>
    <div class="drop-text">Drag & drop file here, or <span onclick="document.getElementById('fileInput').click()">browse</span></div>
    <input type="file" id="fileInput" accept=".xlsx,.xls,.pdf" style="display:none">
    <div class="file-name" id="fileName"></div>
  </div>

  <div class="options">
    <div class="field">
      <label>Distributor</label>
      <select id="distributor">
        <option value="">Auto-detect from filename</option>
        <option value="icedream">Icedream</option>
        <option value="mayyan">Ma'ayan</option>
        <option value="biscotti">Biscotti</option>
        <option value="karfree">Karfree (warehouse stock)</option>
      </select>
    </div>
    <label class="toggle-row">
      <input type="checkbox" id="forceCheck">
      <span class="toggle-label">
        Force re-import
        <small>Overwrite existing data for this period</small>
      </span>
    </label>
  </div>

  <button class="btn" id="uploadBtn" disabled onclick="doUpload()">Upload & Ingest</button>

  <div class="spinner" id="spinner">⏳ Uploading and ingesting data...</div>
  <div class="result" id="result"></div>
  <a href="/" class="back-link">← Back to dashboard</a>
</div>

<script>
  const dropZone = document.getElementById('dropZone');
  const fileInput = document.getElementById('fileInput');
  const fileNameEl = document.getElementById('fileName');
  const uploadBtn = document.getElementById('uploadBtn');
  let selectedFile = null;

  fileInput.addEventListener('change', () => setFile(fileInput.files[0]));

  dropZone.addEventListener('dragover', e => { e.preventDefault(); dropZone.classList.add('drag-over'); });
  dropZone.addEventListener('dragleave', () => dropZone.classList.remove('drag-over'));
  dropZone.addEventListener('drop', e => {
    e.preventDefault();
    dropZone.classList.remove('drag-over');
    setFile(e.dataTransfer.files[0]);
  });

  function setFile(f) {
    if (!f) return;
    selectedFile = f;
    fileNameEl.textContent = '📄 ' + f.name + '  (' + (f.size / 1024).toFixed(0) + ' KB)';
    fileNameEl.style.display = 'block';
    uploadBtn.disabled = false;
  }

  async function doUpload() {
    if (!selectedFile) return;
    const spinner = document.getElementById('spinner');
    const result = document.getElementById('result');

    uploadBtn.disabled = true;
    spinner.style.display = 'block';
    result.style.display = 'none';

    const fd = new FormData();
    fd.append('file', selectedFile, selectedFile.name);
    fd.append('distributor', document.getElementById('distributor').value);
    fd.append('force', document.getElementById('forceCheck').checked ? 'true' : 'false');

    try {
      const resp = await fetch('/upload', { method: 'POST', body: fd });
      const data = await resp.json();

      spinner.style.display = 'none';
      result.style.display = 'block';

      if (data.error) {
        result.className = 'result error';
        result.innerHTML = '<div class="result-title">✗ Error</div>' + escHtml(data.error);
      } else if (data.batches_new === 0 && data.batches_skipped > 0 && !data.weekly_override) {
        // All periods already in DB and no weekly_chart_overrides update either
        result.className = 'result skipped';
        result.innerHTML = '<div class="result-title">⚠ Already ingested</div>' +
          '<table><tr><td>File</td><td>' + escHtml(data.filename) + '</td></tr>' +
          '<tr><td>Skipped</td><td>' + data.batches_skipped + ' batch(es) already in DB</td></tr>' +
          '<tr><td>Tip</td><td>Enable "Force re-import" to overwrite</td></tr></table>';
      } else if (data.batches_new === 0 && data.batches_skipped > 0 && data.weekly_override) {
        // Periods already ingested BUT weekly chart data was updated — normal weekly flow
        var wks = data.weekly_overrides || [data.weekly_override];
        var wkRows = wks.map(function(wk) {
          return '<tr><td>W' + wk.week_num + ' (' + escHtml(wk.label||'') + ')</td><td>' +
            wk.units.toLocaleString() + ' units · ₪' +
            wk.revenue.toLocaleString(undefined,{maximumFractionDigits:0}) + '</td></tr>';
        }).join('');
        result.className = 'result success';
        result.innerHTML = '<div class="result-title">✓ Weekly chart updated</div>' +
          '<table>' +
          '<tr><td>File</td><td>' + escHtml(data.filename) + '</td></tr>' +
          '<tr><td>Distributor</td><td>' + escHtml(data.distributor) + '</td></tr>' +
          wkRows +
          '<tr><td colspan="2" style="color:#a0aec0;font-size:11px;padding-top:6px">Historical periods already in DB — weekly totals updated successfully</td></tr>' +
          '</table>' +
          '<div style="margin-top:12px"><a href="/refresh" style="color:#68d391">→ Refresh dashboard</a></div>';
      } else {
        result.className = 'result success';
        result.innerHTML = '<div class="result-title">✓ Ingestion complete</div>' +
          '<table>' +
          '<tr><td>File</td><td>' + escHtml(data.filename) + '</td></tr>' +
          '<tr><td>Distributor</td><td>' + escHtml(data.distributor) + '</td></tr>' +
          '<tr><td>Type</td><td>' + escHtml(data.type) + '</td></tr>' +
          '<tr><td>Rows ingested</td><td>' + data.rows_processed.toLocaleString() + '</td></tr>' +
          '<tr><td>New batches</td><td>' + data.batches_new + '</td></tr>' +
          (data.batches_skipped ? '<tr><td>Batches skipped</td><td>' + data.batches_skipped + '</td></tr>' : '') +
          '<tr><td>Time</td><td>' + data.elapsed_s.toFixed(1) + 's</td></tr>' +
          '</table>' +
          '<div style="margin-top:12px"><a href="/refresh" style="color:#68d391">→ Refresh dashboard</a></div>';
      }
    } catch (err) {
      spinner.style.display = 'none';
      result.style.display = 'block';
      result.className = 'result error';
      result.innerHTML = '<div class="result-title">✗ Network error</div>' + escHtml(String(err));
    }
    uploadBtn.disabled = false;
  }

  function escHtml(s) {
    return String(s).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');
  }
</script>
</body>
</html>"""


@app.route('/upload', methods=['GET', 'POST'])
def upload():
    """File upload page — ingest a distributor file directly into Cloud SQL."""
    if request.method == 'GET':
        return Response(_UPLOAD_HTML, mimetype='text/html; charset=utf-8')

    # ── POST: handle uploaded file ────────────────────────────────────────────
    import tempfile
    import shutil
    from datetime import datetime
    from pathlib import Path

    file = request.files.get('file')
    if not file or not file.filename:
        return jsonify({'error': 'No file received.'}), 400

    distributor_override = request.form.get('distributor', '').strip() or None
    force = request.form.get('force') == 'true'

    # Save to a temp dir using the original filename (needed for auto-detection)
    tmp_dir = tempfile.mkdtemp(prefix='raito_upload_')
    try:
        safe_name = Path(file.filename).name
        tmp_path = Path(tmp_dir) / safe_name
        file.save(str(tmp_path))

        # Import only what we need — no DB tables required beyond weekly_chart_overrides
        from db.raito_loader import detect_distributor, _copy_to_data_folder, STOCK_PATTERNS

        # Detect distributor
        if distributor_override:
            distributor = distributor_override
            is_stock = (distributor == 'karfree') or any(
                p in safe_name.lower() for p in STOCK_PATTERNS
            )
        else:
            distributor, is_stock = detect_distributor(tmp_path)
            if not distributor:
                return jsonify({'error': (
                    f'Could not auto-detect distributor from filename "{safe_name}". '
                    'Please select a distributor manually.'
                )}), 400

        file_type = 'inventory/stock' if is_stock else 'sales'
        log.info(f"Upload: {safe_name} → {distributor} [{file_type}] force={force}")
        started_at = datetime.now()

        # ── Step 1: Copy file into the data/ subfolder so parsers can find it ──
        # The data/ folder is baked into the Docker image but is writable at runtime.
        # Files copied here persist until the container is replaced (next deploy).
        # On the next dashboard request, parsers.consolidate_data() will pick them up.
        subfolder_map = {
            'icedream': 'icedreams',
            'mayyan': 'mayyan',
            'biscotti': 'biscotti',
            'karfree': 'karfree',
        }
        try:
            subfolder = subfolder_map.get(distributor, distributor)
            dest_dir = _copy_to_data_folder(tmp_path, subfolder)
            log.info(f"  File copied to {dest_dir}/{safe_name}")
        except Exception as e:
            log.error(f"Failed to copy file to data folder: {e}")
            return jsonify({'error': f'File copy failed: {e}'}), 500

        elapsed = (datetime.now() - started_at).total_seconds()

        # ── Step 2: Extract weekly chart data and write to weekly_chart_overrides ─
        # This table is the only SQL table written by uploads. Agents read from here.
        weekly_overrides = []
        if not is_stock and distributor in ('icedream', 'mayyan', 'biscotti'):
            wos = _extract_week_overrides(distributor, tmp_path)
            if wos:
                _wo_ensure_table()
                wo_conn = _md_conn()
                try:
                    wo_cur = wo_conn.cursor()
                    for wo in wos:
                        wo_cur.execute("""
                            INSERT INTO weekly_chart_overrides
                            (distributor, week_num, units, revenue, label, source_file, updated_at)
                            VALUES (%s, %s, %s, %s, %s, %s, NOW())
                            ON CONFLICT (distributor, week_num)
                            DO UPDATE SET
                                units = EXCLUDED.units,
                                revenue = EXCLUDED.revenue,
                                label = EXCLUDED.label,
                                source_file = EXCLUDED.source_file,
                                updated_at = NOW()
                        """, (distributor, wo['week_num'], wo['units'], wo['revenue'], wo['label'], safe_name))
                        log.info(f"Weekly override stored: {distributor} W{wo['week_num']}: {wo['units']} units, ₪{wo['revenue']}")
                    wo_conn.commit()
                    weekly_overrides = wos
                except Exception as e:
                    log.error(f"Failed to store weekly overrides: {e}")
                finally:
                    wo_cur.close()
                    wo_conn.close()

        weekly_override = max(weekly_overrides, key=lambda r: r['week_num']) if weekly_overrides else None

        # ── Step 3: Invalidate dashboard cache ───────────────────────────────────
        # Invalidate cache + start background rebuild so next GET / is fast
        _invalidate_cache()

        response = {
            'filename': safe_name,
            'distributor': distributor,
            'type': file_type,
            'rows_processed': 0,   # kept for UI compat (no DB row count in this flow)
            'batches_new': 1,      # always treat as new so UI shows success
            'batches_skipped': 0,
            'elapsed_s': elapsed,
        }
        if weekly_override:
            response['weekly_override'] = weekly_override
        if weekly_overrides:
            response['weekly_overrides'] = weekly_overrides

        return jsonify(response)

    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


# ══════════════════════════════════════════════════════════════════════════════
# MASTER DATA API  (/api/*)
#
# Backed by a `master_data` table in Cloud SQL (same DB as sales data).
# Each entity (brands, products, customers …) is stored as a JSONB array.
# On first request the table is auto-created and seeded from the Excel file.
# After that all reads and writes go straight to the DB — changes survive
# container restarts and are shared across Cloud Run instances.
#
# Table DDL (auto-created):
#   CREATE TABLE master_data (
#       entity     TEXT PRIMARY KEY,
#       data       JSONB NOT NULL DEFAULT '[]',
#       updated_at TIMESTAMPTZ DEFAULT NOW()
#   );
# ══════════════════════════════════════════════════════════════════════════════

import json as _json

# Primary-key field for each entity (used by PUT / DELETE URL routing)
_MD_PK = {
    'brands':        'key',
    'products':      'sku_key',
    'manufacturers': 'key',
    'distributors':  'key',
    'customers':     'key',
    'logistics':     'product_key',
    'pricing':       '_pk',       # synthetic composite: sku_key::customer::distributor
}

def _pricing_pk(rec: dict) -> str:
    """Generate a composite PK for pricing records."""
    return f"{rec.get('sku_key', '')}::{rec.get('customer', '')}::{rec.get('distributor', '')}"

_MD_ENTITIES = list(_MD_PK.keys())


# ── Price History helpers ────────────────────────────────────────────────────

def _resolve_id(conn, table, key_col, key_val):
    """Resolve a text key to an integer id from a relational table.

    Tries exact match first, then case-insensitive match.
    For customers, also tries the 'key' column if name_en doesn't match.
    """
    if not key_val:
        return None
    cur = conn.cursor()
    # Exact match
    cur.execute(f"SELECT id FROM {table} WHERE {key_col} = %s LIMIT 1", (key_val,))
    row = cur.fetchone()
    if row:
        return row[0]
    # Case-insensitive match
    cur.execute(f"SELECT id FROM {table} WHERE LOWER({key_col}) = LOWER(%s) LIMIT 1", (key_val,))
    row = cur.fetchone()
    return row[0] if row else None


def _price_history_write(record, action='create', old_record=None):
    """Write to price_history when a pricing record is created/updated/deleted.

    Maps MD-tab text keys (sku_key, customer, distributor) to relational integer IDs.
    For 'create': inserts new active price rows.
    For 'update': closes old rows (effective_to=today), inserts new if price changed.
    For 'delete': closes active rows (effective_to=today).
    """
    try:
        conn = _md_conn()
        cur = conn.cursor()
        today = __import__('datetime').date.today()

        # Resolve text keys → integer IDs
        product_id = _resolve_id(conn, 'products', 'sku_key', record.get('sku_key'))
        customer_id = _resolve_id(conn, 'customers', 'name_en', record.get('customer'))
        distributor_id = _resolve_id(conn, 'distributors', 'key', record.get('distributor'))

        if not product_id:
            log.warning("price_history: could not resolve product_id for %s", record.get('sku_key'))
            conn.close()
            return

        if action == 'delete':
            # Close all active rows for this combination
            cur.execute("""
                UPDATE price_history SET effective_to = %s
                WHERE product_id = %s
                  AND (customer_id = %s OR (customer_id IS NULL AND %s IS NULL))
                  AND (distributor_id = %s OR (distributor_id IS NULL AND %s IS NULL))
                  AND effective_to IS NULL
            """, (today, product_id, customer_id, customer_id,
                  distributor_id, distributor_id))

        elif action == 'create':
            # Insert new active rows for sale_price and cost
            for field, ptype in [('sale_price', 'negotiated'), ('cost', 'production_cost')]:
                val = record.get(field)
                if val is not None:
                    try:
                        val = float(val)
                        if val > 0:
                            cur.execute("""
                                INSERT INTO price_history
                                    (product_id, customer_id, distributor_id,
                                     price_ils, effective_from, price_type, source_reference)
                                VALUES (%s, %s, %s, %s, %s, %s, %s)
                            """, (product_id, customer_id, distributor_id,
                                  val, today, ptype, 'md_tab'))
                    except (ValueError, TypeError):
                        pass

        elif action == 'update' and old_record:
            # Check if sale_price or cost actually changed
            for field, ptype in [('sale_price', 'negotiated'), ('cost', 'production_cost')]:
                old_val = old_record.get(field)
                new_val = record.get(field)
                # Normalize for comparison
                try:
                    old_f = float(old_val) if old_val is not None else None
                except (ValueError, TypeError):
                    old_f = None
                try:
                    new_f = float(new_val) if new_val is not None else None
                except (ValueError, TypeError):
                    new_f = None

                if old_f != new_f:
                    # Close old row
                    cur.execute("""
                        UPDATE price_history SET effective_to = %s
                        WHERE product_id = %s
                          AND (customer_id = %s OR (customer_id IS NULL AND %s IS NULL))
                          AND (distributor_id = %s OR (distributor_id IS NULL AND %s IS NULL))
                          AND price_type = %s
                          AND effective_to IS NULL
                    """, (today, product_id, customer_id, customer_id,
                          distributor_id, distributor_id, ptype))
                    # Insert new row
                    if new_f is not None and new_f > 0:
                        cur.execute("""
                            INSERT INTO price_history
                                (product_id, customer_id, distributor_id,
                                 price_ils, effective_from, price_type, source_reference)
                            VALUES (%s, %s, %s, %s, %s, %s, %s)
                        """, (product_id, customer_id, distributor_id,
                              new_f, today, ptype, 'md_tab'))

        conn.commit()
        conn.close()
    except Exception as e:
        log.error("price_history write failed: %s", e)
        try:
            conn.close()
        except Exception:
            pass


def _md_conn():
    """Open a psycopg2 connection using the same DATABASE_URL the rest of the app uses."""
    import psycopg2
    url = os.environ.get('DATABASE_URL', 'postgresql://raito:raito@localhost:5432/raito')
    conn = psycopg2.connect(url)
    conn.autocommit = False
    return conn


def _wo_ensure_table() -> None:
    """Create weekly_chart_overrides table if absent."""
    conn = _md_conn()
    try:
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS weekly_chart_overrides (
                id          SERIAL PRIMARY KEY,
                distributor TEXT NOT NULL,
                week_num    SMALLINT NOT NULL,
                units       INT NOT NULL DEFAULT 0,
                revenue     NUMERIC(12,2) NOT NULL DEFAULT 0,
                label       TEXT,
                source_file TEXT,
                updated_at  TIMESTAMPTZ DEFAULT NOW(),
                UNIQUE(distributor, week_num)
            )
        """)
        conn.commit()
        log.info("weekly_chart_overrides table ensured")
    finally:
        conn.close()


def _md_ensure_table() -> None:
    """Create master_data table if absent; seed from Excel if empty."""
    conn = _md_conn()
    try:
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS master_data (
                entity     TEXT PRIMARY KEY,
                data       JSONB NOT NULL DEFAULT '[]',
                updated_at TIMESTAMPTZ DEFAULT NOW()
            )
        """)
        conn.commit()
        cur.execute("SELECT COUNT(*) FROM master_data")
        if cur.fetchone()[0] == 0:
            _md_seed(cur)
            conn.commit()
            log.info("master_data table seeded from Excel")
    finally:
        conn.close()


def _md_seed(cur) -> None:
    """Populate master_data table from the Excel master-data file."""
    from master_data_parser import parse_master_data
    md = parse_master_data() or {}
    for entity in _MD_ENTITIES:
        cur.execute("""
            INSERT INTO master_data (entity, data, updated_at)
            VALUES (%s, %s, NOW())
            ON CONFLICT (entity) DO UPDATE
                SET data = EXCLUDED.data, updated_at = NOW()
        """, (entity, _json.dumps(md.get(entity, []))))


def _md_read(entity: str) -> list:
    """Fetch one entity's data from Cloud SQL. Returns a Python list."""
    conn = _md_conn()
    try:
        cur = conn.cursor()
        cur.execute("SELECT data FROM master_data WHERE entity = %s", (entity,))
        row = cur.fetchone()
        return row[0] if row else []   # psycopg2 auto-decodes JSONB → Python
    finally:
        conn.close()


def _md_read_all() -> dict:
    """Fetch all entities in one round-trip."""
    conn = _md_conn()
    try:
        cur = conn.cursor()
        cur.execute("SELECT entity, data FROM master_data")
        return {e: d for e, d in cur.fetchall()}
    finally:
        conn.close()


def _md_write(entity: str, data: list) -> None:
    """Upsert one entity's data back to Cloud SQL."""
    conn = _md_conn()
    try:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO master_data (entity, data, updated_at)
            VALUES (%s, %s, NOW())
            ON CONFLICT (entity) DO UPDATE
                SET data = EXCLUDED.data, updated_at = NOW()
        """, (entity, _json.dumps(data)))
        conn.commit()
    finally:
        conn.close()


# ── Audit log (Phase 2) ─────────────────────────────────────────────────────

def _md_ensure_audit_table() -> None:
    """Create master_data_audit table if absent."""
    conn = _md_conn()
    try:
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS master_data_audit (
                id          SERIAL PRIMARY KEY,
                entity      TEXT NOT NULL,
                pk_value    TEXT NOT NULL,
                action      TEXT NOT NULL,
                old_value   JSONB,
                new_value   JSONB,
                actor       TEXT NOT NULL DEFAULT 'anonymous',
                occurred_at TIMESTAMPTZ DEFAULT NOW()
            )
        """)
        cur.execute("""
            CREATE INDEX IF NOT EXISTS idx_audit_entity
            ON master_data_audit (entity, occurred_at DESC)
        """)
        conn.commit()
    finally:
        conn.close()


def _md_audit_log(entity: str, pk_value: str, action: str,
                  old_value: dict | None, new_value: dict | None,
                  actor: str = 'anonymous') -> None:
    """Write one audit row. Called after a successful write."""
    conn = _md_conn()
    try:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO master_data_audit
                (entity, pk_value, action, old_value, new_value, actor, occurred_at)
            VALUES (%s, %s, %s, %s, %s, %s, NOW())
        """, (
            entity, pk_value, action,
            _json.dumps(old_value) if old_value else None,
            _json.dumps(new_value) if new_value else None,
            actor,
        ))
        conn.commit()
    except Exception as exc:
        log.warning("audit_log write failed (non-fatal): %s", exc)
    finally:
        conn.close()


def _md_rebuild_portfolio(md: dict) -> None:
    """Recompute the portfolio matrix from current md state and save to DB.
    Row format matches master_data_parser.py exactly so the JS renderer works:
    headers: ['#','Customer','Customer (EN)','Type','Distributor','Status', ...skus..., 'Active SKUs']
    rows:    [i, name_he, name_en, type, distributor, status, ...price/Pipeline/None..., active_count]
    """
    products = md.get('products', [])
    customers = md.get('customers', [])
    pricing   = md.get('pricing', [])
    active_products = [p for p in products if p.get('status') in ('Active', 'New')]
    if not active_products or not customers:
        portfolio = {'headers': [], 'rows': []}
    else:
        pricing_lookup: dict = {}
        for pr in pricing:
            cust = pr.get('customer', '')
            sku  = pr.get('sku_key', '')
            if cust and sku:
                pricing_lookup[(cust, sku)] = (pr.get('sale_price'), pr.get('status', 'Active'))
        headers = ['#', 'Customer', 'Customer (EN)', 'Type', 'Distributor', 'Status']
        for prod in active_products:
            headers.append(prod.get('name_en') or prod.get('sku_key', ''))
        headers.append('Active SKUs')
        rows = []
        for i, cust in enumerate(sorted(customers, key=lambda c: c.get('name_en', '')), 1):
            name_he  = cust.get('name_he', '')
            name_en  = cust.get('name_en', '')
            cust_key = cust.get('key', '')
            row: list = [i, name_he, name_en,
                         cust.get('type', ''),
                         cust.get('distributor', ''),
                         cust.get('status', '')]
            active_count = 0
            for prod in active_products:
                sku = prod.get('sku_key', '')
                lkp = (pricing_lookup.get((name_he, sku))
                       or pricing_lookup.get((cust_key, sku))
                       or pricing_lookup.get((name_en, sku)))
                if lkp:
                    price, status = lkp
                    if status == 'Pipeline':
                        row.append('Pipeline')
                    elif price:
                        row.append(price)
                        active_count += 1
                    else:
                        row.append(None)
                else:
                    row.append(None)
            row.append(active_count)
            rows.append(row)
        portfolio = {'headers': headers, 'rows': rows}
    # Persist portfolio back to DB
    _md_write('portfolio', portfolio)   # type: ignore[arg-type]
    md['portfolio'] = portfolio


# ── /api/health ───────────────────────────────────────────────────────────────

@app.route('/api/health')
def api_health():
    try:
        _md_ensure_table()
        return jsonify({'status': 'ok', 'source': 'cloud-sql'})
    except Exception as e:
        return jsonify({'status': 'error', 'detail': str(e)}), 500


# ── /api/<entity>  GET (list) ─────────────────────────────────────────────────

@app.route('/api/<string:entity>', methods=['GET'])
def api_list(entity):
    if entity not in _MD_PK:
        return jsonify({'error': f'Unknown entity: {entity}'}), 404
    try:
        _md_ensure_table()
        items = _md_read(entity)
        # Inject synthetic composite PK for pricing records
        if entity == 'pricing':
            for r in items:
                r['_pk'] = _pricing_pk(r)
        return jsonify(items)
    except Exception as e:
        log.error("api_list %s: %s", entity, e)
        return jsonify({'error': str(e)}), 500


# ── Validation import (Phase 2) ──────────────────────────────────────────────

try:
    from db.md_validation import validate_record as _validate
    log.info("Validation module loaded")
except ImportError:
    log.warning("md_validation not found — write routes UNVALIDATED")
    def _validate(entity, record, all_data, action='create', old_record=None):
        return [], []


def _actor() -> str:
    """Return the current admin username from the session, or 'anonymous'."""
    return session.get('md_user', 'anonymous')


# ── /api/cache-status  GET (Phase 4 — poll after writes) ────────────────────

@app.route('/api/cache-status')
def api_cache_status():
    """Return the current state of the dashboard HTML cache."""
    return jsonify({
        'cached': _cached_html is not None,
        'rebuilding': _cache_rebuilding,
        'status': 'ready' if _cached_html is not None else (
            'rebuilding' if _cache_rebuilding else 'stale'),
    })


# ── /api/<entity>  POST (create) ──────────────────────────────────────────────

@app.route('/api/<string:entity>', methods=['POST'])
@require_admin
def api_create(entity):
    if entity not in _MD_PK:
        return jsonify({'error': f'Unknown entity: {entity}'}), 404
    try:
        _md_ensure_table()
        _md_ensure_audit_table()
        record   = request.get_json(force=True) or {}
        # Strip synthetic _pk if the client sent it
        record.pop('_pk', None)
        pk_field = _MD_PK[entity]
        items    = _md_read(entity)

        # Duplicate check: use composite PK for pricing
        if entity == 'pricing':
            new_pk = _pricing_pk(record)
            if any(_pricing_pk(r) == new_pk for r in items):
                return jsonify({'error': f'Duplicate pricing: {new_pk}'}), 409
            pk_val = new_pk
        else:
            pk_val = record.get(pk_field)
            if pk_val and any(r.get(pk_field) == pk_val for r in items):
                return jsonify({'error': f'Duplicate key: {pk_val}'}), 409

        # Validate
        all_data = _md_read_all()
        all_data[entity] = items
        errors, warnings = _validate(entity, record, all_data, action='create')
        if errors:
            return jsonify({'errors': errors}), 422
        if warnings and not request.args.get('confirmed'):
            return jsonify({'warnings': warnings, 'confirm_required': True}), 409

        items.append(record)
        _md_write(entity, items)
        _md_audit_log(entity, str(pk_val or ''), 'create', None, record, _actor())

        if entity == 'pricing':
            _price_history_write(record, action='create')

        if entity in ('products', 'customers', 'pricing'):
            md = _md_read_all()
            _md_rebuild_portfolio(md)

        _invalidate_cache()
        return jsonify(record), 201
    except Exception as e:
        log.error("api_create %s: %s", entity, e)
        return jsonify({'error': str(e)}), 500


# ── /api/<entity>/<pk>  PUT (update) ──────────────────────────────────────────

@app.route('/api/<string:entity>/<path:pk>', methods=['PUT'])
@require_admin
def api_update(entity, pk):
    if entity not in _MD_PK:
        return jsonify({'error': f'Unknown entity: {entity}'}), 404
    try:
        _md_ensure_table()
        _md_ensure_audit_table()
        record   = request.get_json(force=True) or {}
        record.pop('_pk', None)  # strip synthetic PK
        pk_field = _MD_PK[entity]
        items    = _md_read(entity)

        def _item_pk(item):
            """Get the PK value for an item, computing composite PK for pricing."""
            if entity == 'pricing':
                return _pricing_pk(item)
            return str(item.get(pk_field, ''))

        for i, item in enumerate(items):
            if _item_pk(item) == pk:
                old_record = dict(item)
                merged = dict(item)   # start with all old fields
                merged.update(record) # overwrite with submitted fields

                all_data = _md_read_all()
                all_data[entity] = items
                errors, warnings = _validate(
                    entity, merged, all_data,
                    action='update', old_record=item,
                )
                if errors:
                    return jsonify({'errors': errors}), 422
                if warnings and not request.args.get('confirmed'):
                    return jsonify({'warnings': warnings, 'confirm_required': True}), 409

                items[i] = merged
                _md_write(entity, items)
                _md_audit_log(entity, pk, 'update', old_record, record, _actor())
                if entity == 'pricing':
                    _price_history_write(merged, action='update', old_record=old_record)

                if entity in ('products', 'customers', 'pricing'):
                    md = _md_read_all()
                    _md_rebuild_portfolio(md)

                _invalidate_cache()
                return jsonify(record)

        return jsonify({'error': f'Not found: {pk}'}), 404
    except Exception as e:
        log.error("api_update %s/%s: %s", entity, pk, e)
        return jsonify({'error': str(e)}), 500


# ── /api/<entity>/<pk>  DELETE ────────────────────────────────────────────────

@app.route('/api/<string:entity>/<path:pk>', methods=['DELETE'])
@require_admin
def api_delete(entity, pk):
    if entity not in _MD_PK:
        return jsonify({'error': f'Unknown entity: {entity}'}), 404
    try:
        _md_ensure_table()
        _md_ensure_audit_table()
        pk_field = _MD_PK[entity]
        items    = _md_read(entity)

        def _item_pk(r):
            if entity == 'pricing':
                return _pricing_pk(r)
            return str(r.get(pk_field, ''))

        deleted_record = None
        new_items = []
        for r in items:
            if _item_pk(r) == pk:
                deleted_record = r
            else:
                new_items.append(r)

        if deleted_record is None:
            return jsonify({'error': f'Not found: {pk}'}), 404

        all_data = _md_read_all()
        all_data[entity] = items
        errors, _ = _validate(entity, deleted_record, all_data, action='delete')
        if errors:
            return jsonify({'errors': errors}), 422

        _md_write(entity, new_items)
        _md_audit_log(entity, pk, 'delete', deleted_record, None, _actor())
        if entity == 'pricing':
            _price_history_write(deleted_record, action='delete')

        if entity in ('products', 'customers', 'pricing'):
            md = _md_read_all()
            _md_rebuild_portfolio(md)

        _invalidate_cache()
        return jsonify({'deleted': pk})
    except Exception as e:
        log.error("api_delete %s/%s: %s", entity, pk, e)
        return jsonify({'error': str(e)}), 500


# ── /api/pricing/history  GET ─────────────────────────────────────────────────

@app.route('/api/pricing/history', methods=['GET'])
@require_admin
def api_pricing_history():
    """Return price history timeline for a specific pricing entry.

    Query params:
        sku_key     — product SKU key (required)
        customer    — customer name_en (optional)
        distributor — distributor key (optional)
    """
    sku_key = request.args.get('sku_key', '')
    customer = request.args.get('customer', '')
    distributor = request.args.get('distributor', '')

    if not sku_key:
        return jsonify({'error': 'sku_key is required'}), 400

    try:
        conn = _md_conn()
        product_id = _resolve_id(conn, 'products', 'sku_key', sku_key)
        customer_id = _resolve_id(conn, 'customers', 'name_en', customer) if customer else None
        distributor_id = _resolve_id(conn, 'distributors', 'key', distributor) if distributor else None

        if not product_id:
            conn.close()
            return jsonify({'history': [], 'message': f'Product {sku_key} not found in relational table'})

        cur = conn.cursor()
        # Build dynamic WHERE clause
        conditions = ["product_id = %s"]
        params = [product_id]

        if customer_id is not None:
            conditions.append("customer_id = %s")
            params.append(customer_id)
        elif customer:
            # Customer name given but not resolved — return empty
            conn.close()
            return jsonify({'history': [], 'message': f'Customer {customer} not found in relational table'})

        if distributor_id is not None:
            conditions.append("distributor_id = %s")
            params.append(distributor_id)
        elif distributor:
            conn.close()
            return jsonify({'history': [], 'message': f'Distributor {distributor} not found in relational table'})

        where = " AND ".join(conditions)
        cur.execute(f"""
            SELECT ph.id, ph.price_ils, ph.effective_from, ph.effective_to,
                   ph.price_type, ph.source_reference, ph.created_at
            FROM price_history ph
            WHERE {where}
            ORDER BY ph.effective_from DESC, ph.created_at DESC
            LIMIT 100
        """, params)

        rows = cur.fetchall()
        history = []
        for r in rows:
            history.append({
                'id':               r[0],
                'price_ils':        float(r[1]) if r[1] is not None else None,
                'effective_from':   r[2].isoformat() if r[2] else None,
                'effective_to':     r[3].isoformat() if r[3] else None,
                'price_type':       r[4],
                'source_reference': r[5],
                'created_at':       r[6].isoformat() if r[6] else None,
            })

        conn.close()
        return jsonify({'history': history})
    except Exception as e:
        log.error("api_pricing_history: %s", e)
        return jsonify({'error': str(e)}), 500


# ── /api/master-data/upload-excel  POST (bulk import) ────────────────────────

_pending_upload: dict | None = None   # server-side staging for bulk import

@app.route('/api/master-data/upload-excel', methods=['POST'])
@require_admin
def api_md_upload_excel():
    """Bulk import from .xlsx. Step 1: upload file → diff preview.
    Step 2: POST with ?commit=1 → apply changes."""
    global _pending_upload
    try:
        from db.md_excel_roundtrip import parse_upload, diff_preview
    except ImportError:
        from md_excel_roundtrip import parse_upload, diff_preview

    try:
        _md_ensure_table()
        _md_ensure_audit_table()
        commit = request.args.get('commit') == '1'

        if commit:
            uploaded = _pending_upload
            if not uploaded:
                return jsonify({'error': 'No pending upload.'}), 400
            current = _md_read_all()
            actor = _actor()
            entities_written = []
            for entity_key, rows in uploaded.items():
                if entity_key in _MD_PK:
                    old_data = current.get(entity_key, [])
                    _md_write(entity_key, rows)
                    _md_audit_log(entity_key, '*bulk*', 'bulk_upload',
                                  {'count': len(old_data)},
                                  {'count': len(rows)}, actor)
                    entities_written.append(entity_key)
            md = _md_read_all()
            _md_rebuild_portfolio(md)
            _pending_upload = None
            _invalidate_cache()
            return jsonify({'status': 'ok', 'entities_updated': entities_written})

        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded.'}), 400
        f = request.files['file']
        if not f.filename or not f.filename.endswith('.xlsx'):
            return jsonify({'error': 'File must be .xlsx'}), 400
        file_bytes = f.read()
        uploaded = parse_upload(file_bytes)
        current = _md_read_all()
        preview = diff_preview(uploaded, current)
        _pending_upload = uploaded
        return jsonify({'status': 'preview', 'diff': preview})
    except Exception as e:
        log.error("api_md_upload_excel: %s", e)
        return jsonify({'error': str(e)}), 500


# ── /api/master-data/export-excel  GET ───────────────────────────────────────

@app.route('/api/master-data/export-excel', methods=['GET'])
def api_md_export_excel():
    """Download all master data as a formatted .xlsx file."""
    try:
        from db.md_excel_roundtrip import export_xlsx
    except ImportError:
        from md_excel_roundtrip import export_xlsx
    try:
        _md_ensure_table()
        all_data = _md_read_all()
        xlsx_bytes = export_xlsx(all_data)
        return Response(
            xlsx_bytes,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment; filename=Raito_Master_Data_export.xlsx'},
        )
    except Exception as e:
        log.error("api_md_export_excel: %s", e)
        return jsonify({'error': str(e)}), 500


# ── /api/pricing/bulk-apply  POST ────────────────────────────────────────────

@app.route('/api/pricing/bulk-apply', methods=['POST'])
@require_admin
def api_pricing_bulk_apply():
    """Bulk price change. Body: {filter, operation, value, field, commit}."""
    try:
        from db.md_excel_roundtrip import bulk_price_preview, apply_bulk_price
    except ImportError:
        from md_excel_roundtrip import bulk_price_preview, apply_bulk_price
    try:
        _md_ensure_table()
        _md_ensure_audit_table()
        body = request.get_json(force=True) or {}
        filter_spec = body.get('filter', {})
        operation = body.get('operation', '')
        value = body.get('value', 0)
        field = body.get('field', 'sale_price')
        commit = body.get('commit', False)

        if operation not in ('pct', 'absolute', 'set'):
            return jsonify({'error': "operation must be 'pct', 'set', or 'absolute'"}), 400
        if not isinstance(value, (int, float)):
            return jsonify({'error': 'value must be a number'}), 400
        if field not in ('sale_price', 'cost'):
            return jsonify({'error': "field must be 'sale_price' or 'cost'"}), 400

        pricing = _md_read('pricing')
        changes = bulk_price_preview(pricing, filter_spec, operation, value, field)

        if not commit:
            return jsonify({'status': 'preview', 'changes': changes, 'count': len(changes)})

        updated_pricing = apply_bulk_price(pricing, changes)
        _md_write('pricing', updated_pricing)
        _md_audit_log('pricing', '*bulk*', 'bulk_price',
                      {'operation': operation, 'value': value, 'field': field,
                       'filter': filter_spec},
                      {'affected_count': len(changes)}, _actor())
        md = _md_read_all()
        _md_rebuild_portfolio(md)
        _invalidate_cache()
        return jsonify({'status': 'ok', 'applied': len(changes)})
    except Exception as e:
        log.error("api_pricing_bulk_apply: %s", e)
        return jsonify({'error': str(e)}), 500


# ── /api/audit-log  GET ─────────────────────────────────────────────────────

@app.route('/api/audit-log', methods=['GET'])
@require_admin
def api_audit_log():
    """Return recent audit log entries. Supports ?entity=X&limit=N."""
    try:
        _md_ensure_audit_table()
        entity = request.args.get('entity')
        limit = int(request.args.get('limit', 50))
        conn = _md_conn()
        try:
            cur = conn.cursor()
            if entity:
                cur.execute("""
                    SELECT id, entity, pk_value, action, old_value, new_value, actor, occurred_at
                    FROM master_data_audit WHERE entity = %s
                    ORDER BY occurred_at DESC LIMIT %s
                """, (entity, limit))
            else:
                cur.execute("""
                    SELECT id, entity, pk_value, action, old_value, new_value, actor, occurred_at
                    FROM master_data_audit
                    ORDER BY occurred_at DESC LIMIT %s
                """, (limit,))
            rows = []
            for r in cur.fetchall():
                rows.append({
                    'id': r[0], 'entity': r[1], 'pk_value': r[2],
                    'action': r[3], 'old_value': r[4], 'new_value': r[5],
                    'actor': r[6],
                    'occurred_at': r[7].isoformat() if r[7] else None,
                })
            return jsonify(rows)
        finally:
            conn.close()
    except Exception as e:
        log.error("api_audit_log: %s", e)
        return jsonify({'error': str(e)}), 500


# ── /api/portfolio  GET ───────────────────────────────────────────────────────

@app.route('/api/portfolio', methods=['GET'])
def api_portfolio():
    try:
        _md_ensure_table()
        # Always rebuild fresh — computation is fast and avoids stale cache
        md = _md_read_all()
        _md_rebuild_portfolio(md)
        return jsonify(md.get('portfolio', {}))
    except Exception as e:
        log.error("api_portfolio: %s", e)
        return jsonify({'error': str(e)}), 500


# ── /api/lookup/<entity>  (FK dropdowns) ──────────────────────────────────────

@app.route('/api/lookup/<string:entity>', methods=['GET'])
def api_lookup(entity):
    try:
        _md_ensure_table()
        items = _md_read(entity)
        # Return full records so the JS SCHEMAS' valKey/labelKey work directly
        return jsonify(items)
    except Exception as e:
        log.error("api_lookup %s: %s", entity, e)
        return jsonify({'error': str(e)}), 500


# ── /api/normalize  (one-time data cleanup) ───────────────────────────────────

@app.route('/api/normalize', methods=['POST'])
@require_admin
def api_normalize():
    """Normalize FK references across all entities to use proper keys.

    Fixes seed data that stored display names (e.g. 'Icedream') instead of
    keys (e.g. 'icedreams') for distributor references. Also deduplicates
    customer records and removes orphan test data.
    """
    try:
        _md_ensure_table()
        md = _md_read_all()
        changes = []

        # Build lookup maps
        dists = md.get('distributors', [])
        custs = md.get('customers', [])

        # Distributor: display name → key
        dist_name_to_key = {}
        for d in dists:
            k = d.get('key', '')
            n = d.get('name', '')
            dist_name_to_key[n] = k
            # Also map common short forms
            for short in [n.split('(')[0].strip(), n.split(' ')[0].strip()]:
                if short:
                    dist_name_to_key[short] = k
        # Known aliases
        dist_name_to_key['Icedream'] = 'icedreams'
        dist_name_to_key["Ma'ayan"] = 'mayyan_froz'
        dist_name_to_key['Biscotti'] = 'biscotti'
        dist_name_to_key['Ma\'ayan'] = 'mayyan_froz'

        # Customer: display name → key
        cust_name_to_key = {}
        cust_key_map = {}
        for c in custs:
            ck = c.get('key', '')
            cust_key_map[ck] = c
            for field in ('name_en', 'name_he', 'key'):
                v = c.get(field, '')
                if v:
                    cust_name_to_key[v] = ck

        # ── Normalize customers.distributor → key ──
        for c in custs:
            old_dist = c.get('distributor', '')
            if old_dist and old_dist not in [d.get('key') for d in dists]:
                new_dist = dist_name_to_key.get(old_dist)
                if new_dist:
                    c['distributor'] = new_dist
                    changes.append(f"customer {c.get('name_en','?')}: distributor {old_dist} → {new_dist}")

        # ── Deduplicate customers by key ──
        seen_keys = {}
        deduped_custs = []
        for c in custs:
            ck = c.get('key', '')
            if ck in seen_keys:
                changes.append(f"customer duplicate removed: {c.get('name_en','?')} (key={ck})")
            else:
                seen_keys[ck] = True
                deduped_custs.append(c)
        custs = deduped_custs

        # ── Normalize pricing.distributor → key, pricing.customer → key ──
        pricing = md.get('pricing', [])
        for p in pricing:
            old_dist = p.get('distributor', '')
            if old_dist and old_dist not in [d.get('key') for d in dists]:
                new_dist = dist_name_to_key.get(old_dist)
                if new_dist:
                    p['distributor'] = new_dist
                    changes.append(f"pricing {p.get('sku_key','?')}/{p.get('customer','?')}: distributor {old_dist} → {new_dist}")

            old_cust = p.get('customer', '')
            if old_cust and old_cust not in [c.get('key') for c in custs]:
                new_cust = cust_name_to_key.get(old_cust)
                if new_cust:
                    p['customer'] = new_cust
                    changes.append(f"pricing {p.get('sku_key','?')}/{old_cust}: customer → {new_cust}")

        # ── Write back ──
        if changes:
            _md_write('customers', custs)
            _md_write('pricing', pricing)
            _md_audit_log('*system*', '*normalize*', 'normalize',
                          {'changes_count': len(changes)},
                          {'changes': changes[:50]}, _actor())

        return jsonify({
            'status': 'ok',
            'changes': changes,
            'total_changes': len(changes),
        })
    except Exception as e:
        log.error("api_normalize: %s", e, exc_info=True)
        return jsonify({'error': str(e)}), 500


# ── /api/weekly-overrides  (dynamic chart data) ───────────────────────────────

@app.route('/api/weekly-overrides', methods=['GET'])
def api_weekly_overrides_get():
    """Fetch all weekly chart override data grouped by distributor."""
    try:
        _wo_ensure_table()
        conn = _md_conn()
        try:
            cur = conn.cursor()
            cur.execute("""
                SELECT distributor, week_num, units, revenue
                FROM weekly_chart_overrides
                ORDER BY distributor, week_num
            """)
            rows = cur.fetchall()
        finally:
            conn.close()

        # Structure as {distributor: {week: {units, revenue}}}
        result = {}
        for dist, wk, units, revenue in rows:
            if dist not in result:
                result[dist] = {}
            result[dist][str(wk)] = {
                'units': int(units),
                'revenue': float(revenue)
            }
        return jsonify(result)
    except Exception as e:
        log.error("api_weekly_overrides_get: %s", e)
        return jsonify({'error': str(e)}), 500


@app.route('/api/weekly-overrides', methods=['POST'])
def api_weekly_overrides_post():
    """Store or update a weekly override from an uploaded file."""
    try:
        _wo_ensure_table()
        data = request.get_json(force=True) or {}
        distributor = data.get('distributor', '').strip()
        week_num = data.get('week_num')
        units = data.get('units', 0)
        revenue = data.get('revenue', 0)
        label = data.get('label', '')
        source_file = data.get('source_file', '')

        if not distributor or week_num is None:
            return jsonify({'error': 'Missing distributor or week_num'}), 400

        conn = _md_conn()
        try:
            cur = conn.cursor()
            cur.execute("""
                INSERT INTO weekly_chart_overrides
                (distributor, week_num, units, revenue, label, source_file, updated_at)
                VALUES (%s, %s, %s, %s, %s, %s, NOW())
                ON CONFLICT (distributor, week_num)
                DO UPDATE SET
                    units = EXCLUDED.units,
                    revenue = EXCLUDED.revenue,
                    label = EXCLUDED.label,
                    source_file = EXCLUDED.source_file,
                    updated_at = NOW()
            """, (distributor, week_num, units, revenue, label, source_file))
            conn.commit()
        finally:
            conn.close()

        return jsonify({'ok': True, 'week_num': week_num})
    except Exception as e:
        log.error("api_weekly_overrides_post: %s", e)
        return jsonify({'error': str(e)}), 500


# ── /api/rebuild  (regenerate dashboard from current DB state) ────────────────

@app.route('/api/rebuild', methods=['POST'])
def api_rebuild():
    """Recompute portfolio, invalidate dashboard cache + background rebuild."""
    try:
        _md_ensure_table()
        md = _md_read_all()
        _md_rebuild_portfolio(md)
        try:
            from db.database_manager import reset_cache
            reset_cache()
        except Exception:
            pass
        _invalidate_cache()
        log.info("API rebuild triggered — background rebuild started")
        return jsonify({'status': 'rebuilding',
                        'message': 'Dashboard rebuilding in background'})
    except Exception as e:
        log.error("api_rebuild: %s", e)
        return jsonify({'error': str(e)}), 500


# ── Standalone mode (dev server) ──────────────────────────────────────────

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 8080))
    log.info(f"Starting Flask dev server on :{port}")
    app.run(host='0.0.0.0', port=port, debug=True)
