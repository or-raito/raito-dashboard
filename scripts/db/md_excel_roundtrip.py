#!/usr/bin/env python3
"""
Excel round-trip for Master Data — upload (diff preview + commit) and export.

Phase 3 SSOT migration (2026-04-19):
  - parse_upload(file_bytes) → {entity: [rows...]}
  - diff_preview(uploaded, current) → {entity: {added:[], changed:[], removed:[]}}
  - export_xlsx(all_data) → BytesIO containing the .xlsx

Used by db_dashboard.py routes:
  POST /api/master-data/upload-excel
  GET  /api/master-data/export-excel
"""

from __future__ import annotations

import io
import logging
from typing import Optional

log = logging.getLogger(__name__)

# Column layouts per sheet — matches master_data_parser.py export format exactly
SHEET_COLUMNS: dict[str, list[dict]] = {
    'Brands': [
        {'key': 'key',         'label': 'Brand Key'},
        {'key': 'name',        'label': 'Brand Name'},
        {'key': 'category',    'label': 'Category'},
        {'key': 'status',      'label': 'Status'},
        {'key': 'launch_date', 'label': 'Launch Date'},
        {'key': 'owner',       'label': 'Owner'},
        {'key': 'notes',       'label': 'Notes'},
    ],
    'Products': [
        {'key': 'sku_key',      'label': 'SKU Key'},
        {'key': 'barcode',      'label': 'Barcode'},
        {'key': 'name_he',      'label': 'Name HE'},
        {'key': 'name_en',      'label': 'Name EN'},
        {'key': 'brand',         'label': 'Brand Key'},
        {'key': 'category',     'label': 'Category'},
        {'key': 'status',       'label': 'Status'},
        {'key': 'launch_date',  'label': 'Launch Date'},
        {'key': 'manufacturer', 'label': 'Manufacturer'},
        {'key': 'cost',         'label': 'Cost'},
    ],
    'Manufacturers': [
        {'key': 'key',           'label': 'Key'},
        {'key': 'name',          'label': 'Name'},
        {'key': 'products',      'label': 'Products'},
        {'key': 'contact',       'label': 'Contact'},
        {'key': 'location',      'label': 'Location'},
        {'key': 'lead_time',     'label': 'Lead Time'},
        {'key': 'moq',           'label': 'MOQ'},
        {'key': 'payment_terms', 'label': 'Payment Terms'},
        {'key': 'notes',         'label': 'Notes'},
    ],
    'Distributors': [
        {'key': 'key',           'label': 'Key'},
        {'key': 'name',          'label': 'Name'},
        {'key': 'products',      'label': 'Products'},
        {'key': 'commission_pct','label': 'Commission'},
        {'key': 'report_format', 'label': 'Report Format'},
        {'key': 'report_freq',   'label': 'Report Freq'},
        {'key': 'contact',       'label': 'Contact'},
        {'key': 'notes',         'label': 'Notes'},
    ],
    'Customers': [
        {'key': 'key',        'label': 'Customer Key'},
        {'key': 'name_he',    'label': 'Name HE'},
        {'key': 'name_en',    'label': 'Name EN'},
        {'key': 'type',       'label': 'Type'},
        {'key': 'distributor','label': 'Distributor'},
        {'key': 'chain',      'label': 'Chain/Group'},
        {'key': 'status',     'label': 'Status'},
        {'key': 'contact',    'label': 'Contact'},
        {'key': 'phone',      'label': 'Phone'},
        {'key': 'notes',      'label': 'Notes'},
    ],
    'Logistics': [
        {'key': 'product_key',       'label': 'Product Key'},
        {'key': 'product_name',      'label': 'Product Name'},
        {'key': 'storage_type',      'label': 'Storage Type'},
        {'key': 'temp',              'label': 'Temp'},
        {'key': 'units_per_carton',  'label': 'Units/Carton'},
        {'key': 'cartons_per_pallet','label': 'Cartons/Pallet'},
        {'key': 'units_per_pallet',  'label': 'Units/Pallet'},
        {'key': 'pallet_divisor',    'label': 'Pallet Divisor'},
        {'key': 'warehouse',         'label': 'Warehouse'},
        {'key': 'notes',             'label': 'Notes'},
    ],
    'Pricing': [
        {'key': 'barcode',        'label': 'Barcode'},
        {'key': 'sku_key',        'label': 'SKU Key'},
        {'key': 'name_en',        'label': 'Name EN'},
        {'key': 'name_he',        'label': 'Name HE'},
        {'key': 'customer',       'label': 'Customer'},
        {'key': 'distributor',    'label': 'Distributor'},
        {'key': 'commission_pct', 'label': 'Commission'},
        {'key': 'sale_price',     'label': 'Sale Price'},
        {'key': 'cost',           'label': 'Cost'},
        {'key': 'gross_margin',   'label': 'Gross Margin'},
    ],
}

# Sheet name → (entity key in master_data, primary key field)
SHEET_TO_ENTITY: dict[str, tuple[str, str]] = {
    'Brands':        ('brands',        'key'),
    'Products':      ('products',      'sku_key'),
    'Manufacturers': ('manufacturers', 'key'),
    'Distributors':  ('distributors',  'key'),
    'Customers':     ('customers',     'key'),
    'Logistics':     ('logistics',     'product_key'),
    'Pricing':       ('pricing',       '_composite'),  # sku_key::customer::distributor
}


# ═══════════════════════════════════════════════════════════════════════════════
# Upload: parse Excel bytes → dict of entity rows
# ═══════════════════════════════════════════════════════════════════════════════

def parse_upload(file_bytes: bytes) -> dict[str, list[dict]]:
    """Parse an uploaded .xlsx file into {entity_key: [row_dicts...]}.

    Handles the 'export' format (row 1 = headers, data from row 2).
    """
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    result: dict[str, list[dict]] = {}

    for sheet_name, (entity_key, _pk) in SHEET_TO_ENTITY.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        cols = SHEET_COLUMNS.get(sheet_name, [])
        if not cols:
            continue

        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Skip empty rows (check first cell)
            if row[0] is None and (len(row) < 2 or row[1] is None):
                continue
            record = {}
            for i, col_def in enumerate(cols):
                val = row[i] if i < len(row) else None
                # Convert to string for text fields, keep numbers as-is
                if val is None:
                    record[col_def['key']] = ''
                elif isinstance(val, (int, float)):
                    record[col_def['key']] = val
                else:
                    record[col_def['key']] = str(val).strip()
            rows.append(record)
        result[entity_key] = rows

    return result


# ═══════════════════════════════════════════════════════════════════════════════
# Diff preview: compare uploaded data against current DB data
# ═══════════════════════════════════════════════════════════════════════════════

def _row_pk(row: dict, pk_field: str) -> str:
    """Compute PK for a row. Handles composite pricing PK."""
    if pk_field == '_composite':
        parts = [
            str(row.get('sku_key', '')).strip(),
            str(row.get('customer', '')).strip(),
            str(row.get('distributor', '')).strip(),
        ]
        return '::'.join(parts) if any(parts) else ''
    return str(row.get(pk_field, '')).strip()


def diff_preview(
    uploaded: dict[str, list[dict]],
    current: dict[str, list[dict]],
) -> dict[str, dict]:
    """Compare uploaded data to current DB data.

    Returns: {
        entity: {
            'added':   [records not in current],
            'changed': [{'old': {...}, 'new': {...}, 'fields': ['field1',...]}],
            'removed': [records in current but not in uploaded],
            'unchanged': int,
        }
    }
    """
    result = {}

    for sheet_name, (entity_key, pk_field) in SHEET_TO_ENTITY.items():
        up_rows = uploaded.get(entity_key, [])
        cur_rows = current.get(entity_key, [])

        if not up_rows and not cur_rows:
            continue

        # Index current by PK
        cur_by_pk: dict[str, dict] = {}
        for r in cur_rows:
            pk = _row_pk(r, pk_field)
            if pk:
                cur_by_pk[pk] = r

        added = []
        changed = []
        seen_pks = set()
        unchanged = 0

        for r in up_rows:
            pk = _row_pk(r, pk_field)
            if not pk:
                continue
            seen_pks.add(pk)

            if pk not in cur_by_pk:
                added.append(r)
            else:
                # Compare fields
                old = cur_by_pk[pk]
                diff_fields = []
                for col_def in SHEET_COLUMNS.get(sheet_name, []):
                    key = col_def['key']
                    old_val = _normalize(old.get(key))
                    new_val = _normalize(r.get(key))
                    if old_val != new_val:
                        diff_fields.append(key)
                if diff_fields:
                    changed.append({
                        'old': old,
                        'new': r,
                        'fields': diff_fields,
                    })
                else:
                    unchanged += 1

        removed = [r for pk_val, r in cur_by_pk.items() if pk_val not in seen_pks]

        result[entity_key] = {
            'added': added,
            'changed': changed,
            'removed': removed,
            'unchanged': unchanged,
        }

    return result


def _normalize(val) -> str:
    """Normalize a value for comparison (string, lowercase, stripped)."""
    if val is None:
        return ''
    if isinstance(val, float):
        # Avoid float precision issues: 0.30000000000000004 → '0.3'
        if val == int(val):
            return str(int(val))
        return f'{val:.6f}'.rstrip('0').rstrip('.')
    return str(val).strip()


# ═══════════════════════════════════════════════════════════════════════════════
# Bulk price operations
# ═══════════════════════════════════════════════════════════════════════════════

def bulk_price_preview(
    pricing: list[dict],
    filter_spec: dict,
    operation: str,
    value: float,
    field: str = 'sale_price',
) -> list[dict]:
    """Preview bulk price changes.

    Args:
        pricing:    Current pricing rows
        filter_spec: {distributor?: str, customer?: str, sku_key?: str, brand?: str}
        operation:  'pct' (percentage) or 'absolute' (fixed amount)
        value:      The change amount (e.g., 5 for +5% or +₪5)
        field:      'sale_price' or 'cost'

    Returns:
        List of {record, old_value, new_value} for affected rows.
    """
    results = []
    for row in pricing:
        if not _matches_filter(row, filter_spec):
            continue
        old_val = row.get(field)
        if old_val is None:
            continue
        try:
            old_val = float(old_val)
        except (ValueError, TypeError):
            continue

        if operation == 'pct':
            new_val = round(old_val * (1 + value / 100), 2)
        elif operation == 'absolute':
            new_val = round(old_val + value, 2)
        else:
            continue

        if new_val != old_val:
            results.append({
                'barcode': row.get('barcode', ''),
                'sku_key': row.get('sku_key', ''),
                'customer': row.get('customer', ''),
                'distributor': row.get('distributor', ''),
                'field': field,
                'old_value': old_val,
                'new_value': new_val,
            })

    return results


def _pricing_composite(row: dict) -> str:
    """Build composite key for pricing lookup."""
    return f"{row.get('sku_key', '')}::{row.get('customer', '')}::{row.get('distributor', '')}"


def apply_bulk_price(
    pricing: list[dict],
    changes: list[dict],
) -> list[dict]:
    """Apply bulk price changes to pricing data. Returns the updated list."""
    # Build a lookup of changes by composite key
    change_map = {}
    for c in changes:
        change_map[_pricing_composite(c)] = c

    updated = []
    for row in pricing:
        pk = _pricing_composite(row)
        if pk in change_map:
            c = change_map[pk]
            row = dict(row)  # copy
            row[c['field']] = c['new_value']
            # Recompute margins
            sp = row.get('sale_price')
            cost = row.get('cost')
            if sp is not None and cost is not None:
                try:
                    sp = float(sp)
                    cost = float(cost)
                    row['gross_profit'] = round(sp - cost, 2)
                    row['gross_margin'] = round((sp - cost) / sp, 4) if sp > 0 else None
                except (ValueError, TypeError):
                    pass
        updated.append(row)

    return updated


def _matches_filter(row: dict, f: dict) -> bool:
    """Check if a pricing row matches the filter specification."""
    for key in ('distributor', 'customer', 'sku_key'):
        fval = f.get(key, '')
        if fval and row.get(key, '') != fval:
            return False
    # Brand filter requires product lookup — skip if not provided
    # (caller should pre-filter or the UI should resolve brand → SKUs)
    return True


# ═══════════════════════════════════════════════════════════════════════════════
# Export: generate .xlsx from current DB data
# ═══════════════════════════════════════════════════════════════════════════════

def export_xlsx(all_data: dict[str, list]) -> bytes:
    """Generate a .xlsx file from all master data entities.

    Returns the file as bytes (ready for send_file).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='E8EAF6', end_color='E8EAF6', fill_type='solid')
    header_border = Border(bottom=Side(style='thin', color='9E9E9E'))
    header_align = Alignment(horizontal='center', vertical='center')

    for sheet_name, (entity_key, _pk) in SHEET_TO_ENTITY.items():
        rows = all_data.get(entity_key, [])
        cols = SHEET_COLUMNS.get(sheet_name, [])
        if not cols:
            continue

        ws = wb.create_sheet(title=sheet_name)

        # Headers
        for ci, col_def in enumerate(cols, 1):
            cell = ws.cell(row=1, column=ci, value=col_def['label'])
            cell.font = header_font
            cell.fill = header_fill
            cell.border = header_border
            cell.alignment = header_align

        # Data rows
        for ri, record in enumerate(rows, 2):
            for ci, col_def in enumerate(cols, 1):
                val = record.get(col_def['key'], '')
                if val is None:
                    val = ''
                ws.cell(row=ri, column=ci, value=val)

        # Auto-width columns (approximate)
        for ci, col_def in enumerate(cols, 1):
            max_len = len(col_def['label'])
            for ri in range(2, min(len(rows) + 2, 50)):  # sample first 50 rows
                cell_val = str(ws.cell(row=ri, column=ci).value or '')
                max_len = max(max_len, len(cell_val))
            ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = min(max_len + 4, 40)

        # Freeze header row
        ws.freeze_panes = 'A2'

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
