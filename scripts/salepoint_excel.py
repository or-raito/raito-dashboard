#!/usr/bin/env python3
"""
Raito Sale Points Excel Export — Deep-dive workbook with customer and branch data.
Generates multi-sheet Excel matching the reference raito_salepoints_deep_dive format.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config import PRODUCT_SHORT
from registry import CUSTOMER_NAMES_EN
from pricing_engine import get_b2b_price_safe
from business_logic import compute_status, compute_trend, compute_ordering_pattern

# ── Constants ──────────────────────────────────────────────────────────────────

from config import _MONTH_REGISTRY
MONTHS = [m[0] for m in _MONTH_REGISTRY]
MONTH_SHORT = {m[0]: m[2].split()[0] for m in _MONTH_REGISTRY}
FLAVORS = ['chocolate', 'vanilla', 'mango', 'pistachio', 'dream_cake', 'dream_cake_2']

# Ma'ayan customer name normalization (typo/variant → canonical Hebrew, then EN)
_MAAYAN_CUSTOMER_NORM = {
    'פז יילו': 'פז ילו',
    'פז  ילו': 'פז ילו',
}

# Icedream customer prefix → canonical customer name
_ICE_CUSTOMER_PREFIXES = [
    ('וולט מרקט', 'וולט מרקט'),
    ('וואלט', 'וולט מרקט'),
    ('וולט', 'וולט מרקט'),
    ('דומינוס', 'דומינוס'),
    ('גוד פארם', 'גוד פארם'),
    ('חוות נעמי', 'חוות נעמי'),
    ('נוי השדה', 'נוי השדה'),
    ('ינגו', 'ינגו'),
    ('כרמלה', 'כרמלה'),
    ('פוט לוקר', 'פוט לוקר'),
]

def _en(name):
    """Translate customer name to English."""
    return CUSTOMER_NAMES_EN.get(name, CUSTOMER_NAMES_EN.get(name.strip(), name))

# Styles
HDR_FILL = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
ALT_FILL = PatternFill(start_color='F8F9F9', end_color='F8F9F9', fill_type='solid')
NO_FILL = PatternFill(fill_type=None)
CHURN_FILL = PatternFill(start_color='FDEDEC', end_color='FDEDEC', fill_type='solid')
REACT_FILL = PatternFill(start_color='EAFAF1', end_color='EAFAF1', fill_type='solid')

HDR_FONT = Font(bold=True, color='FFFFFF')
GREEN_BOLD = Font(bold=True, color='27AE60')
RED_BOLD = Font(bold=True, color='E74C3C')
BLUE_BOLD = Font(bold=True, color='2980B9')
NAVY_TITLE = Font(bold=True, size=14, color='2C3E50')
GREY_BOLD = Font(bold=True, color='64748B')

CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center')
RIGHT = Alignment(horizontal='right', vertical='center')


# ── Customer helpers ───────────────────────────────────────────────────────────

def _get_ice_customer(name: str) -> str:
    """Extract canonical Icedream customer name from a branch name string."""
    n = name.strip()
    for prefix, customer in _ICE_CUSTOMER_PREFIXES:
        if n.startswith(prefix):
            return _en(customer)
    return _en(n)  # fallback: use full name as customer


def _safe_sheet_name(name: str) -> str:
    """Sanitize a string for use as an Excel sheet name."""
    result = name[:31]
    for ch in [':', '*', '?', '/', '\\', '[', ']']:
        result = result.replace(ch, '_')
    return result


# ── Status / trend helpers ─────────────────────────────────────────────────────
# Delegated to business_logic.py (SSOT). Local wrappers for backward compat.

def _compute_status(*monthly_values) -> str:
    """Delegate to business_logic.compute_status (SSOT)."""
    return compute_status(*monthly_values)


def _compute_trend_fraction(first, last):
    """First→Last trend as a fraction for Excel formatting."""
    from business_logic import compute_trend_fraction
    return compute_trend_fraction(first, last)


def _compute_ordering_pattern(*monthly_values) -> str:
    """Delegate to business_logic.compute_ordering_pattern (SSOT)."""
    return compute_ordering_pattern(*monthly_values)


# ── Data extraction ────────────────────────────────────────────────────────────

def _empty_group(name: str, distributor: str) -> dict:
    return {
        'name': name,
        'distributor': distributor,
        'salepoints': {},
        'months_data': {m: 0 for m in MONTHS},
        'flavor_breakdown': {f: 0 for f in FLAVORS},
        'total_units': 0,
        'total_revenue': 0.0,
    }


def _empty_sp(name: str) -> dict:
    return {
        'name': name,
        'months_units': {m: 0 for m in MONTHS},
        'flavor_breakdown': {f: 0 for f in FLAVORS},
        'total_units': 0,
        'total_revenue': 0.0,
    }


def _extract_salepoint_data(data: dict) -> dict:
    """Build grouped sale-point data from consolidated parser output."""
    group_map = {}  # (group_name, distributor) → group dict

    for month in MONTHS:
        if month not in data.get('monthly_data', {}):
            continue
        md = data['monthly_data'][month]

        # ── Icedream customers ─────────────────────────────────────────────
        for cust_name, products in md.get('icedreams_customers', {}).items():
            customer = _get_ice_customer(cust_name)
            key = (customer, 'Icedream')
            if key not in group_map:
                group_map[key] = _empty_group(customer, 'Icedream')
            grp = group_map[key]

            sp_name = cust_name
            if sp_name not in grp['salepoints']:
                grp['salepoints'][sp_name] = _empty_sp(sp_name)
            sp = grp['salepoints'][sp_name]

            for flavor, pdata in products.items():
                if flavor not in FLAVORS:
                    continue
                if isinstance(pdata, dict):
                    units = pdata.get('units', 0) or 0
                    value = pdata.get('value', 0) or 0
                else:
                    units = int(pdata) if pdata else 0
                    value = units * get_b2b_price_safe(flavor)

                sp['months_units'][month] += units
                sp['total_units'] += units
                sp['total_revenue'] += value
                sp['flavor_breakdown'][flavor] += units

                grp['months_data'][month] += units
                grp['total_units'] += units
                grp['total_revenue'] += value
                grp['flavor_breakdown'][flavor] += units

        # ── Ma'ayan accounts ───────────────────────────────────────────────
        for (customer_name, acct_name), products in md.get('mayyan_accounts', {}).items():
            customer_name = _MAAYAN_CUSTOMER_NORM.get(customer_name, customer_name)
            customer_name = _en(customer_name)
            key = (customer_name, "Ma'ayan")
            if key not in group_map:
                group_map[key] = _empty_group(customer_name, "Ma'ayan")
            grp = group_map[key]

            sp_name = acct_name if acct_name else f'{customer_name} (Branch)'
            if sp_name not in grp['salepoints']:
                grp['salepoints'][sp_name] = _empty_sp(sp_name)
            sp = grp['salepoints'][sp_name]

            for flavor, prod_data in products.items():
                if flavor not in FLAVORS:
                    continue
                units = int(prod_data.get('units', 0)) if isinstance(prod_data, dict) else int(prod_data or 0)
                value = prod_data.get('value', 0) if isinstance(prod_data, dict) else 0

                sp['months_units'][month] += units
                sp['total_units'] += units
                sp['total_revenue'] += value
                sp['flavor_breakdown'][flavor] += units

                grp['months_data'][month] += units
                grp['total_units'] += units
                grp['total_revenue'] += value
                grp['flavor_breakdown'][flavor] += units

        # ── Biscotti customers (direct distribution) ─────────────────────
        for branch_name, products in md.get('biscotti_customers', {}).items():
            customer = 'Biscotti Customer'
            key = (customer, 'Biscotti')
            if key not in group_map:
                group_map[key] = _empty_group(customer, 'Biscotti')
            grp = group_map[key]

            sp_name = branch_name
            if sp_name not in grp['salepoints']:
                grp['salepoints'][sp_name] = _empty_sp(sp_name)
            sp = grp['salepoints'][sp_name]

            for flavor, pdata in products.items():
                if flavor not in FLAVORS:
                    continue
                if isinstance(pdata, dict):
                    units = pdata.get('units', 0) or 0
                    value = pdata.get('value', 0) or 0
                else:
                    units = int(pdata) if pdata else 0
                    value = units * get_b2b_price_safe(flavor)

                sp['months_units'][month] += units
                sp['total_units'] += units
                sp['total_revenue'] += value
                sp['flavor_breakdown'][flavor] += units

                grp['months_data'][month] += units
                grp['total_units'] += units
                grp['total_revenue'] += value
                grp['flavor_breakdown'][flavor] += units

    # ── Compute derived fields ─────────────────────────────────────────────
    groups = sorted(group_map.values(), key=lambda g: g['total_units'], reverse=True)

    for grp in groups:
        month_vals = [grp['months_data'].get(m, 0) for m in MONTHS]

        grp['dec_trend'] = _compute_trend_fraction(month_vals[0], month_vals[-1]) if len(month_vals) >= 2 else None
        grp['ordering_pattern'] = _compute_ordering_pattern(*month_vals)

        # Active SP count per month
        grp['active_per_month'] = {}
        for m in MONTHS:
            grp['active_per_month'][m] = sum(
                1 for sp in grp['salepoints'].values() if sp['months_units'].get(m, 0) > 0
            )

        # Avg units per active SP for last month
        last_m = MONTHS[-1]
        last_active = grp['active_per_month'].get(last_m, 0)
        last_units = grp['months_data'].get(last_m, 0)
        grp['avg_units_per_sp_feb'] = (last_units // last_active) if last_active > 0 else 0

        # Per-SP derived fields
        sps = sorted(grp['salepoints'].values(), key=lambda s: s['total_units'], reverse=True)
        grp['salepoints_sorted'] = sps
        for sp in sps:
            sp_vals = [sp['months_units'].get(m, 0) for m in MONTHS]
            sp['status'] = _compute_status(*sp_vals)
            sp['trend'] = _compute_trend_fraction(sp_vals[0], sp_vals[-1]) if len(sp_vals) >= 2 else None
            sp['months_active'] = sum(1 for v in sp_vals if v > 0)

    return {
        'groups': groups,
        'total_groups': len(groups),
        'total_salepoints': sum(len(g['salepoints']) for g in groups),
        'total_units': sum(g['total_units'] for g in groups),
        'total_revenue': sum(g['total_revenue'] for g in groups),
    }


# ── Style helpers ──────────────────────────────────────────────────────────────

def _apply_header_row(ws, row: int, headers: list):
    """Write a navy header row."""
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = CENTER


def _row_fill(data_index: int):
    """Alternating fill: odd rows (1-indexed) get light grey, even rows nothing."""
    return ALT_FILL if data_index % 2 == 1 else NO_FILL


def _status_font(status: str):
    if status == 'Active':
        return GREEN_BOLD
    elif status == 'Churned':
        return RED_BOLD
    elif status == 'Reactivated':
        return BLUE_BOLD
    return None  # default font for Mar gap / New


def _status_row_fill(status: str, alt_fill):
    if status == 'Churned':
        return CHURN_FILL
    elif status == 'Reactivated':
        return REACT_FILL
    return alt_fill


def _trend_font(trend):
    if trend is None:
        return None
    return GREEN_BOLD if trend >= 0 else RED_BOLD


def _apply_data_row(ws, row: int, values: list, num_fmts: dict = None,
                    row_fill=None, fonts: dict = None):
    """Write a data row with optional per-column formatting."""
    for col_idx, value in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = value
        if row_fill:
            cell.fill = row_fill
        if num_fmts and col_idx in num_fmts:
            cell.number_format = num_fmts[col_idx]
        if fonts and col_idx in fonts and fonts[col_idx] is not None:
            cell.font = fonts[col_idx]


# ── Sheet builders ─────────────────────────────────────────────────────────────

def _create_summary_sheet(wb: Workbook, sp: dict):
    """Customer Summary — 17 cols, header row 1, data rows 2+."""
    ws = wb.create_sheet('Customer Summary', 0)

    n_m = len(MONTHS)
    m_short = [MONTH_SHORT[m] for m in MONTHS]
    last_short = m_short[-1] if m_short else 'Last'

    headers = (
        ['#', 'Distributor', 'Customer', 'Total Sale Points']
        + [f'{s} Active' for s in m_short]
        + ['Total Units']
        + [f'{s} Units' for s in m_short]
        + ['Total Revenue', f'Avg Units/Point ({last_short})',
           f'{m_short[0]}→{m_short[-1]} Trend', 'Ordering Pattern']
    )
    _apply_header_row(ws, 1, headers)

    # Number formats by column index (1-based)
    # Cols: #(1), Dist(2), Cust(3), TotSP(4), [n_m active], TotUnits, [n_m units], TotRev, Avg, Trend, Pattern
    num_fmts = {}
    base_units = 4 + n_m + 1  # first unit column
    for j in range(n_m):
        num_fmts[base_units + j] = '#,##0'
    num_fmts[base_units - 1] = '#,##0'  # Total Units
    num_fmts[base_units + n_m] = '₪#,##0'  # Total Revenue
    num_fmts[base_units + n_m + 1] = '#,##0'  # Avg
    num_fmts[base_units + n_m + 2] = '0%'  # Trend

    for i, grp in enumerate(sp['groups'], 1):
        row = i + 1
        fill = _row_fill(i)

        month_units = [grp['months_data'].get(m, 0) for m in MONTHS]
        trend = grp.get('dec_trend')

        values = (
            [i, grp['distributor'], grp['name'], len(grp['salepoints'])]
            + [grp['active_per_month'].get(m, 0) for m in MONTHS]
            + [grp['total_units']]
            + month_units
            + [grp['total_revenue'], grp.get('avg_units_per_sp_feb', 0),
               trend, grp.get('ordering_pattern', 'Consistent')]
        )

        trend_col = 4 + n_m + 1 + n_m + 2  # Trend column index
        fonts = {trend_col: _trend_font(trend)}
        _apply_data_row(ws, row, values, num_fmts=num_fmts, row_fill=fill, fonts=fonts)

    # Column widths
    total_cols = 4 + n_m + 1 + n_m + 3
    for ci in range(1, total_cols + 1):
        col_letter = get_column_letter(ci)
        if ci <= 3:
            ws.column_dimensions[col_letter].width = [4, 10, 25][ci - 1]
        elif ci == 4:
            ws.column_dimensions[col_letter].width = 14
        else:
            ws.column_dimensions[col_letter].width = 12

    ws.auto_filter.ref = f'A1:{get_column_letter(total_cols)}{len(sp["groups"]) + 1}'


def _create_all_salepoints_sheet(wb: Workbook, sp: dict):
    """All Sale Points — 17 cols, header row 1, data rows 2+."""
    ws = wb.create_sheet('All Sale Points', 1)

    m_short = [MONTH_SHORT[m] for m in MONTHS]
    n_m = len(MONTHS)
    headers = (
        ['Distributor', 'Customer', 'Sale Point Name']
        + [f'{s} Units' for s in m_short]
        + ['Total Units', 'Total Revenue', 'Months Active',
           f'{m_short[0]}→{m_short[-1]} Change', 'Status',
           'Chocolate', 'Vanilla', 'Mango', 'Pistachio', 'Dream Cake']
    )
    _apply_header_row(ws, 1, headers)

    num_fmts = {}
    for j in range(n_m):
        num_fmts[4 + j] = '#,##0'
    num_fmts[4 + n_m] = '#,##0'  # Total Units
    num_fmts[4 + n_m + 1] = '₪#,##0'  # Total Revenue
    num_fmts[4 + n_m + 3] = '0%'  # Trend
    for j in range(5):
        num_fmts[4 + n_m + 5 + j] = '#,##0'  # Flavor columns

    data_row = 2
    data_idx = 1
    for grp in sp['groups']:
        for spoint in grp.get('salepoints_sorted', sorted(grp['salepoints'].values(),
                                                           key=lambda s: s['total_units'],
                                                           reverse=True)):
            status = spoint.get('status', 'Active')
            trend = spoint.get('trend')
            alt = _row_fill(data_idx)
            fill = _status_row_fill(status, alt)

            sp_month_vals = [spoint['months_units'].get(m, 0) for m in MONTHS]

            values = (
                [grp['distributor'], grp['name'], spoint['name']]
                + sp_month_vals
                + [spoint['total_units'], spoint['total_revenue'],
                   spoint.get('months_active', 0), trend, status,
                   spoint['flavor_breakdown'].get('chocolate', 0),
                   spoint['flavor_breakdown'].get('vanilla', 0),
                   spoint['flavor_breakdown'].get('mango', 0),
                   spoint['flavor_breakdown'].get('pistachio', 0),
                   spoint['flavor_breakdown'].get('dream_cake', 0) + spoint['flavor_breakdown'].get('dream_cake_2', 0)]
            )

            sf = _status_font(status)
            trend_col = 4 + n_m + 3  # Trend column
            status_col = 4 + n_m + 4  # Status column
            fonts = {trend_col: _trend_font(trend), status_col: sf}
            _apply_data_row(ws, data_row, values, num_fmts=num_fmts,
                            row_fill=fill, fonts=fonts)

            data_row += 1
            data_idx += 1

    # Column widths
    widths = {'A': 10, 'B': 20, 'C': 55,
              'D': 10, 'E': 10, 'F': 10, 'G': 10,
              'H': 12, 'I': 13, 'J': 11, 'K': 13, 'L': 18,
              'M': 10, 'N': 10, 'O': 10, 'P': 10, 'Q': 10}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.auto_filter.ref = f'A1:Q{data_row - 1}'


def _create_group_sheet(wb: Workbook, grp: dict):
    """Per-customer-group sheet: stats header (rows 1-3), headers (row 5), data (row 6+)."""
    sheet_name = _safe_sheet_name(grp['name'])
    ws = wb.create_sheet(sheet_name)

    m_short = [MONTH_SHORT[m] for m in MONTHS]
    n_m = len(MONTHS)
    active = grp.get('active_per_month', {})
    active_str = ' → '.join(f'{s}:{active.get(m, 0)}' for m, s in zip(MONTHS, m_short))

    # Row 1: Title (merged A1:F1)
    title = f"{grp['name']} ({grp['distributor']})"
    ws['A1'] = title
    ws['A1'].font = NAVY_TITLE
    ws.merge_cells('A1:F1')

    # Row 2: Stats
    ws['A2'] = 'Total Sale Points:'
    ws['A2'].font = Font(bold=True, size=10)
    ws['B2'] = len(grp['salepoints'])
    ws['C2'] = 'Total Units:'
    ws['C2'].font = Font(bold=True, size=10)
    ws['D2'] = grp['total_units']
    ws['E2'] = 'Total Revenue:'
    ws['E2'].font = Font(bold=True, size=10)
    ws['F2'] = grp['total_revenue']
    ws['F2'].number_format = '₪#,##0'

    # Row 3: Active points progression
    ws['A3'] = 'Active Points:'
    ws['A3'].font = Font(bold=True, size=10)
    ws['B3'] = active_str

    # Row 5: Column headers
    headers = (
        ['#', 'Sale Point']
        + m_short
        + ['Total', 'Months Active', 'Trend', 'Status',
           'Choc', 'Van', 'Mango', 'Pist', 'DC']
    )
    _apply_header_row(ws, 5, headers)

    # Number formats
    num_fmts = {}
    for j in range(n_m):
        num_fmts[3 + j] = '#,##0'
    num_fmts[3 + n_m] = '#,##0'  # Total
    num_fmts[3 + n_m + 2] = '0%'  # Trend
    for j in range(5):
        num_fmts[3 + n_m + 4 + j] = '#,##0'  # Flavor columns

    sps = grp.get('salepoints_sorted', sorted(grp['salepoints'].values(),
                                               key=lambda s: s['total_units'],
                                               reverse=True))
    data_idx = 1
    for idx, sp_item in enumerate(sps, 1):
        row = idx + 5
        status = sp_item.get('status', 'Active')
        trend = sp_item.get('trend')
        alt = _row_fill(data_idx)
        fill = _status_row_fill(status, alt)

        sp_month_vals = [sp_item['months_units'].get(m, 0) for m in MONTHS]

        values = (
            [idx, sp_item['name']]
            + sp_month_vals
            + [sp_item['total_units'], sp_item.get('months_active', 0),
               trend, status,
               sp_item['flavor_breakdown'].get('chocolate', 0),
               sp_item['flavor_breakdown'].get('vanilla', 0),
               sp_item['flavor_breakdown'].get('mango', 0),
               sp_item['flavor_breakdown'].get('pistachio', 0),
               sp_item['flavor_breakdown'].get('dream_cake', 0) + sp_item['flavor_breakdown'].get('dream_cake_2', 0)]
        )

        sf = _status_font(status)
        trend_col = 3 + n_m + 2  # Trend
        status_col = 3 + n_m + 3  # Status
        fonts = {trend_col: _trend_font(trend), status_col: sf}
        _apply_data_row(ws, row, values, num_fmts=num_fmts, row_fill=fill, fonts=fonts)

        data_idx += 1

    # Column widths
    total_cols = 2 + n_m + 8
    for ci in range(1, total_cols + 1):
        col_letter = get_column_letter(ci)
        if ci == 1:
            ws.column_dimensions[col_letter].width = 4
        elif ci == 2:
            ws.column_dimensions[col_letter].width = 55
        else:
            ws.column_dimensions[col_letter].width = 10


# ── Public API ─────────────────────────────────────────────────────────────────

def generate_salepoint_excel(data: dict, output_path):
    """
    Generate comprehensive Excel workbook with sale points data.

    Args:
        data: Consolidated data from consolidate_data()
        output_path: Path to write the Excel file
    """
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    sp = _extract_salepoint_data(data)

    _create_summary_sheet(wb, sp)
    _create_all_salepoints_sheet(wb, sp)

    # Per-group sheets (all groups, ordered by total units)
    for grp in sp['groups']:
        _create_group_sheet(wb, grp)

    wb.save(str(output_path))
    print(f'Sale Points Excel saved: {output_path}')


if __name__ == '__main__':
    import sys
    import io
    from pathlib import Path

    sys.path.insert(0, str(Path(__file__).parent))
    old_stdout = sys.stdout
    sys.stdout = io.StringIO()
    from parsers import consolidate_data
    data = consolidate_data()
    sys.stdout = old_stdout

    output = Path('/tmp/test_salepoint.xlsx')
    generate_salepoint_excel(data, output)
    print(f'Generated: {output}')
