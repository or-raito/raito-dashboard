#!/usr/bin/env python3
"""
Raito Master Data Parser
Reads Raito_Master_Data.xlsx (and the export copy if present) into structured
dicts for the unified dashboard.

Two source files are used:
  EXPORT_PATH  — Raito_Master_Data_export.xlsx  (user-edited via dashboard Save)
                 Clean format: row 1 = headers, data row 2+, no # column.
                 Contains: Brands, Products, Manufacturers, Distributors,
                           Customers, Logistics, Pricing
  ORIGINAL_PATH — Raito_Master_Data.xlsx  (master source)
                 Used for: Portfolio, Config (not exported by dashboard)
                 Also used as fallback for all sheets if export doesn't exist.

Column layouts (export format):
  Brands:        Brand Key(0) Name(1) Category(2) Status(3) Launch Date(4) Owner(5) Notes(6)
  Products:      SKU Key(0) Barcode(1) Name HE(2) Name EN(3) Brand Key(4) Category(5)
                 Status(6) Launch Date(7) Manufacturer(8) Cost(9)
  Manufacturers: Key(0) Name(1) Products(2) Contact(3) Location(4) Lead Time(5)
                 MOQ(6) Payment Terms(7) Notes(8)
  Distributors:  Key(0) Name(1) Products(2) Commission(3) Report Format(4)
                 Report Freq(5) Contact(6) Notes(7)
  Customers:     Customer Key(0) Name HE(1) Name EN(2) Type(3) Distributor(4)
                 Chain/Group(5) Status(6) Contact(7) Phone(8) Notes(9)
  Logistics:     Product Key(0) Product Name(1) Storage Type(2) Temp(3)
                 Units/Carton(4) Cartons/Pallet(5) Units/Pallet(6)
                 Pallet Divisor(7) Warehouse(8) Notes(9)
  Pricing:       Barcode(0) SKU Key(1) Name EN(2) Name HE(3) Customer(4)
                 Distributor(5) Commission(6) Sale Price(7) Cost(8) Gross Margin(9)

  Portfolio (original only):
                 #(0) Customer HEB(1) EN(2) Type(3) Distributor(4) Status(5)
                 then product EN columns (6+), Active SKUs (last)
  Config (original only):
                 Parameter(0) Value(1)
"""

from pathlib import Path
from openpyxl import load_workbook
from config import DATA_DIR

ORIGINAL_PATH = DATA_DIR / 'master data' / 'Raito_Master_Data.xlsx'
EXPORT_PATH   = DATA_DIR / 'master data' / 'Raito_Master_Data_export.xlsx'


def _str(v):
    if v is None:
        return ''
    return str(v).strip()


def _num(v):
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def _rows(ws, min_row=2):
    return list(ws.iter_rows(min_row=min_row, values_only=True))


def parse_master_data():
    """Parse Master Data Excel into structured dicts for dashboard rendering."""

    # ── Choose source files ──────────────────────────────────────────────────
    use_export = EXPORT_PATH.exists()
    src = EXPORT_PATH if use_export else ORIGINAL_PATH

    if not src.exists():
        print(f"Master Data not found: {src}")
        return None

    wb_src = load_workbook(src, data_only=True)

    # Always open the original for Portfolio and Config
    wb_orig = None
    if ORIGINAL_PATH.exists():
        wb_orig = load_workbook(ORIGINAL_PATH, data_only=True)

    result = {}

    # ── Brands ──────────────────────────────────────────────────────────────
    if 'Brands' in wb_src.sheetnames:
        result['brands'] = []
        for r in _rows(wb_src['Brands']):
            if not r[0]:
                continue
            result['brands'].append({
                'key':         _str(r[0]),
                'name':        _str(r[1]),
                'category':    _str(r[2]),
                'status':      _str(r[3]),
                'launch_date': _str(r[4]),
                'owner':       _str(r[5]),
                'notes':       _str(r[6]) if len(r) > 6 else '',
            })

    # ── Products ─────────────────────────────────────────────────────────────
    if 'Products' in wb_src.sheetnames:
        result['products'] = []
        for r in _rows(wb_src['Products']):
            if not r[0]:
                continue
            result['products'].append({
                'sku_key':      _str(r[0]),
                'barcode':      _str(r[1]),
                'name_he':      _str(r[2]),
                'name_en':      _str(r[3]),
                'brand':        _str(r[4]),
                'category':     _str(r[5]),
                'status':       _str(r[6]),
                'launch_date':  _str(r[7]),
                'manufacturer': _str(r[8]) if len(r) > 8 else '',
                'cost':         _num(r[9]) if len(r) > 9 else None,
            })

    # ── Manufacturers ─────────────────────────────────────────────────────────
    if 'Manufacturers' in wb_src.sheetnames:
        result['manufacturers'] = []
        for r in _rows(wb_src['Manufacturers']):
            if not r[0]:
                continue
            result['manufacturers'].append({
                'key':           _str(r[0]),
                'name':          _str(r[1]),
                'products':      _str(r[2]),
                'contact':       _str(r[3]),
                'location':      _str(r[4]),
                'lead_time':     _str(r[5]) if len(r) > 5 else '',
                'moq':           _str(r[6]) if len(r) > 6 else '',
                'payment_terms': _str(r[7]) if len(r) > 7 else '',
                'notes':         _str(r[8]) if len(r) > 8 else '',
            })

    # ── Distributors ──────────────────────────────────────────────────────────
    if 'Distributors' in wb_src.sheetnames:
        result['distributors'] = []
        for r in _rows(wb_src['Distributors']):
            if not r[0]:
                continue
            result['distributors'].append({
                'key':           _str(r[0]),
                'name':          _str(r[1]),
                'products':      _str(r[2]),
                'commission_pct': _num(r[3]) if len(r) > 3 else None,
                'report_format': _str(r[4]) if len(r) > 4 else '',
                'report_freq':   _str(r[5]) if len(r) > 5 else '',
                'contact':       _str(r[6]) if len(r) > 6 else '',
                'notes':         _str(r[7]) if len(r) > 7 else '',
            })

    # ── Customers ─────────────────────────────────────────────────────────────
    # Export: Customer Key at col 0 (no leading # column)
    # Original: # at col 0, Customer Key at col 1 (min_row=5)
    if 'Customers' in wb_src.sheetnames:
        result['customers'] = []
        if use_export:
            for r in _rows(wb_src['Customers'], min_row=2):
                if not r[0]:
                    continue
                result['customers'].append({
                    'key':        _str(r[0]),
                    'name_he':    _str(r[1]),
                    'name_en':    _str(r[2]) if len(r) > 2 else '',
                    'type':       _str(r[3]) if len(r) > 3 else '',
                    'distributor':_str(r[4]) if len(r) > 4 else '',
                    'chain':      _str(r[5]) if len(r) > 5 else '',
                    'status':     _str(r[6]) if len(r) > 6 else '',
                    'contact':    _str(r[7]) if len(r) > 7 else '',
                    'phone':      _str(r[8]) if len(r) > 8 else '',
                    'notes':      _str(r[9]) if len(r) > 9 else '',
                })
        else:
            # Original format: rows 1-3 title/legend, row 4 headers, data row 5+
            for r in _rows(wb_src['Customers'], min_row=5):
                if not r[1]:
                    continue
                result['customers'].append({
                    'key':        _str(r[1]),
                    'name_he':    _str(r[2]),
                    'name_en':    _str(r[3]) if len(r) > 3 else '',
                    'type':       _str(r[4]) if len(r) > 4 else '',
                    'distributor':_str(r[5]) if len(r) > 5 else '',
                    'chain':      _str(r[6]) if len(r) > 6 else '',
                    'status':     _str(r[7]) if len(r) > 7 else '',
                    'contact':    _str(r[8]) if len(r) > 8 else '',
                    'phone':      _str(r[9]) if len(r) > 9 else '',
                    'notes':      _str(r[10]) if len(r) > 10 else '',
                })

    # ── Logistics ─────────────────────────────────────────────────────────────
    if 'Logistics' in wb_src.sheetnames:
        result['logistics'] = []
        for r in _rows(wb_src['Logistics']):
            if not r[0]:
                continue
            result['logistics'].append({
                'product_key':        _str(r[0]),
                'product_name':       _str(r[1]) if len(r) > 1 else '',
                'storage_type':       _str(r[2]) if len(r) > 2 else '',
                'temp':               _str(r[3]) if len(r) > 3 else '',
                'units_per_carton':   r[4] if len(r) > 4 else None,
                'cartons_per_pallet': r[5] if len(r) > 5 else None,
                'units_per_pallet':   r[6] if len(r) > 6 else None,
                'pallet_divisor':     r[7] if len(r) > 7 else None,
                'warehouse':          _str(r[8]) if len(r) > 8 else '',
                'notes':              _str(r[9]) if len(r) > 9 else '',
            })

    # ── Pricing ───────────────────────────────────────────────────────────────
    # Export: Barcode(0) SKU Key(1) Name EN(2) Name HE(3) Customer(4)
    #         Distributor(5) Commission(6) Sale Price(7) Cost(8) Gross Margin(9)
    # Original: #(0) Barcode(1) SKU Key(2) … min_row=4
    if 'Pricing' in wb_src.sheetnames:
        result['pricing'] = []
        if use_export:
            for r in _rows(wb_src['Pricing'], min_row=2):
                if not r[1]:   # SKU Key at col 1
                    continue
                gm = _num(r[9]) if len(r) > 9 else None
                # Compute op_margin from gross_margin and commission if available
                commission = _num(r[6]) if len(r) > 6 else None
                sale_price = _num(r[7]) if len(r) > 7 else None
                cost       = _num(r[8]) if len(r) > 8 else None
                op_margin  = None
                if gm is not None and commission is not None and sale_price and sale_price > 0:
                    dist_comm_val = commission * sale_price if commission else 0
                    op_profit = (sale_price - cost - dist_comm_val) if cost else None
                    if op_profit is not None:
                        op_margin = op_profit / sale_price
                result['pricing'].append({
                    'barcode':         _str(r[0]),
                    'sku_key':         _str(r[1]),
                    'name_en':         _str(r[2]) if len(r) > 2 else '',
                    'name_he':         _str(r[3]) if len(r) > 3 else '',
                    'customer':        _str(r[4]) if len(r) > 4 else '',
                    'distributor':     _str(r[5]) if len(r) > 5 else '',
                    'commission_pct':  commission,
                    'sale_price':      sale_price,
                    'cost':            cost,
                    'gross_profit':    (sale_price - cost) if sale_price and cost else None,
                    'gross_margin':    gm,
                    'dist_commission': None,
                    'op_profit':       None,
                    'op_margin':       op_margin,
                    'status':          'Active',   # export doesn't carry status; default Active
                    'last_updated':    '',
                })
        else:
            # Original format: rows 1-2 title, row 3 headers, data row 4+
            for r in _rows(wb_src['Pricing'], min_row=4):
                if not r[2]:
                    continue
                result['pricing'].append({
                    'barcode':         _str(r[1]) if len(r) > 1 else '',
                    'sku_key':         _str(r[2]),
                    'name_en':         _str(r[3]) if len(r) > 3 else '',
                    'name_he':         _str(r[4]) if len(r) > 4 else '',
                    'customer':        _str(r[5]) if len(r) > 5 else '',
                    'distributor':     _str(r[6]) if len(r) > 6 else '',
                    'commission_pct':  _num(r[7]) if len(r) > 7 else None,
                    'sale_price':      _num(r[8]) if len(r) > 8 else None,
                    'cost':            _num(r[9]) if len(r) > 9 else None,
                    'gross_profit':    _num(r[10]) if len(r) > 10 else None,
                    'gross_margin':    _num(r[11]) if len(r) > 11 else None,
                    'dist_commission': _num(r[12]) if len(r) > 12 else None,
                    'op_profit':       _num(r[13]) if len(r) > 13 else None,
                    'op_margin':       _num(r[14]) if len(r) > 14 else None,
                    'status':          _str(r[15]) if len(r) > 15 else '',
                    'last_updated':    _str(r[16]) if len(r) > 16 else '',
                })

    # ── Normalize pricing: resolve customer names + fill missing product names ─
    if 'customers' in result and 'pricing' in result:
        # Customer name → English lookup
        _to_en = {}
        for c in result['customers']:
            en = c.get('name_en', '')
            if en:
                _to_en[c.get('name_he', '')] = en
                _to_en[c.get('key', '')]     = en
                _to_en[en]                   = en   # identity
        for pr in result['pricing']:
            raw = pr.get('customer', '')
            if raw in _to_en:
                pr['customer'] = _to_en[raw]

    if 'products' in result and 'pricing' in result:
        # Fill missing name_en / name_he from products table by sku_key
        _sku_lookup = {}
        for p in result['products']:
            _sku_lookup[p.get('sku_key', '')] = p
        for pr in result['pricing']:
            sku = pr.get('sku_key', '')
            prod = _sku_lookup.get(sku)
            if prod:
                if not pr.get('name_en'):
                    pr['name_en'] = prod.get('name_en', '')
                if not pr.get('name_he'):
                    pr['name_he'] = prod.get('name_he', '')

    # ── Portfolio — dynamically generated from customers + products + pricing ─
    # Build a pivot: rows = customers, columns = active products, cells = price/Pipeline/empty
    if 'products' in result and 'customers' in result and 'pricing' in result:
        active_products = [p for p in result['products']
                           if p.get('status') in ('Active', 'New')]

        # Build pricing lookup keyed by BOTH customer name_he AND customer key
        # Pricing rows may use either display names (Hebrew) or keys (English)
        pricing_lookup = {}   # (identifier, sku_key) → (sale_price, status)
        for pr in result['pricing']:
            cust_name = pr.get('customer', '')
            sku = pr.get('sku_key', '')
            if cust_name and sku:
                pricing_lookup[(cust_name, sku)] = (
                    pr.get('sale_price'),
                    pr.get('status', 'Active')
                )

        headers = ['#', 'Customer', 'Customer (EN)', 'Type', 'Distributor', 'Status']
        for prod in active_products:
            headers.append(prod.get('name_en', prod.get('sku_key', '')))
        headers.append('Active SKUs')

        rows = []
        sorted_customers = sorted(result['customers'],
                                  key=lambda c: c.get('name_en', ''))
        for i, cust in enumerate(sorted_customers, 1):
            name_he = cust.get('name_he', '')
            name_en = cust.get('name_en', '')
            cust_key = cust.get('key', '')
            row = [i, name_he, name_en,
                   cust.get('type', ''),
                   cust.get('distributor', ''),
                   cust.get('status', '')]
            active_count = 0
            for prod in active_products:
                sku = prod.get('sku_key', '')
                # Try matching by name_he first, then by customer key, then name_en
                lookup = (pricing_lookup.get((name_he, sku))
                          or pricing_lookup.get((cust_key, sku))
                          or pricing_lookup.get((name_en, sku)))
                if lookup:
                    price, status = lookup
                    if status == 'Pipeline':
                        row.append('Pipeline')
                    elif price:
                        row.append(price)
                        active_count += 1
                    else:
                        row.append(None)
                else:
                    row.append(None)
            row.append(active_count)
            rows.append(row)

        result['portfolio'] = {'headers': headers, 'rows': rows}

    # ── Config — always from original ────────────────────────────────────────
    wb_cfg = wb_orig or wb_src
    if 'Config' in wb_cfg.sheetnames:
        result['config'] = []
        for r in _rows(wb_cfg['Config']):
            if not r[0]:
                continue
            result['config'].append({
                'parameter': _str(r[0]),
                'value':     _str(r[1]) if len(r) > 1 else '',
            })

    wb_src.close()
    if wb_orig:
        wb_orig.close()

    src_label = 'export' if use_export else 'original'
    print(f"  Master data loaded from {src_label} ({src.name})")
    return result


if __name__ == '__main__':
    md = parse_master_data()
    if md:
        for key, val in md.items():
            if isinstance(val, list):
                print(f"  {key}: {len(val)} records")
                if val:
                    print(f"    sample: {list(val[0].items())[:4]}")
            elif isinstance(val, dict):
                print(f"  {key}: {len(val.get('rows', []))} rows | headers: {val.get('headers', [])[:6]}")
